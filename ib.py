from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import time
import datetime
import os.path
from shutil import rmtree
import sys
import logging
import platform
from netapp_systems import *
from my_logging import print_to_log
import pickle
import pandas as pd
import xlrd

logging.basicConfig(filename='ib_scrub.log', filemode='w', level=logging.DEBUG)
debug_it = 0


def get_column_letter_new(column_index):
    """Convert a column number into a column letter (3 -> 'C')

    Right shift the column col_idx by 26 to find column letters in reverse
    order.  These numbers are 1-based, and can be converted to ASCII
    ordinals by adding 64.

    """
    # these indicies correspond to A -> ZZZ and include all allowed
    # columns
    if not 1 <= column_index <= 18278:
        msg = 'Column index out of bounds: %s' % column_index
        raise ValueError(msg)
    letters = []
    while column_index > 0:
        column_index, remainder = divmod(column_index, 26)
        # check for exact division and borrow if needed
        if remainder == 0:
            remainder = 26
            column_index -= 1
        letters.append(chr(remainder+64))
    return ''.join(reversed(letters))

column_heading_row = 8
last_data_column = 30
sc_last_data_column = 19
data_sheet_name = "Sheet1"
ib_report_name_cell = "A7"
sc_report_name_cell = "A6"

company_text = "Serial Number Owner Company"
site_text = "Installed At Site Name"
group_text = "Group Name"
solution_text = "Solution"
serial_text = "Serial Number"
hostname_text = "System Name"
cluster_name_text = "Cluster Name"
cluster_serial_text = "Cluster Serial Number"
cluster_uuid_text = "Cluster UUID"
os_version_text = "OS Version"
asup_status_text = "ASUP Status"
declined_text = "ASUP Declined Reason"
asup_date_text = "ASUP Last Generate Date"
product_family_text = "Installed At Product Family"
platform_text = "Installed At Platform"
controller_eos_date_text = "Controller EOS Date"
pvr_flag_text = "EOS PVR Flag"
pvr_date_text = "PVR End Date"
first_eos_date_text = "First EOS Date"
first_eos_part_text = "First EOS Part"
age_text = "System Age in Years"
ha_pair_flag_text = "HA Pair Flag"
service_level_text = "Service Level"
entitlement_status_text = "Entitlement Status"
contact_name_text = "Primary Contact Name"
contact_number_text = "Primary Contact Number"
contact_email_text = "Primary Contact Email"
raw_tb_text = "Raw TB"
num_shelves_text = "# of Shelves"
num_disks_text = "# of Disks"
warranty_end_date_text = "Warranty End Date"
service_type_text = "Service Type"
response_profile_text = "Response Profile"
service_contract_id_text = "Service Contract ID"
contract_end_date_text = "Contract End Date"
contract_extended_date_text = "Contract Extended Date"
months_till_expire_text = "Months Till Expire"
contract_status_text = "Contract Status"
# move to sc section hw_service_level_status_text = "HW Service Level Status"
nrd_service_contract_id_text = "NRD Service Contract ID"
nrd_contract_end_date_text = "NRD Contract End Date"
nrd_contract_extended_date_text = "NRD Contract Extended Date"
nrd_months_till_expire_text = "NRD Months Till Expire"
nrd_hw_service_level_status_text = "NRD HW Service Level Status"
account_name_text = "Account Name"
end_customer_text = "End Customer NAGP"
customer_geo_text = "Installed At Customer Geo"
customer_country_text = "Installed At Customer Country"
customer_city_text = "Installed At Customer City"
original_so_text = "Original Sales Order Number"
shipped_date_text = "Shipped Date"
software_service_end_date_text = "Software Service End Date"
reseller_text = "Reseller Company On Original Order"
service_contracts_report_filename = "Service%20Contract%20Expiring_Service%20Contract%20Expiring.xlsx"
short_s_c_filename = "sc.xlsx"
installed_base_report_filename = "Installed%20Base%20Details_Installed%20Base%20Details.xlsx"
short_i_b_filename = "ib.xlsx"
score_report_filesname = "score.xlsx"
sam_report_filename = "sam.xlsx"
installed_base_hostname_list_filename = "ib.txt"
missing_hostname_filename = "serial_to_hostname.txt"
companies_to_ignore_filename = "companies_to_ignore.txt"
serials_to_ignore_filename = "serials_to_ignore.txt"
hostnames_to_ignore_filename = "hostnames_to_ignore.txt"
groups_to_ignore_filename = "groups_to_ignore.txt"
sites_to_ignore_filename = "sites_to_ignore.txt"
ib_notes_filename = "ib_notes.txt"
override_filename = "overrides.txt"
solidfire_tag_text= "Solidfire Service Tag Number"


def check_input_files():
    if os.path.isfile(installed_base_report_filename):
        if os.path.isfile(short_i_b_filename):
            if os.path.isfile("old_" + short_i_b_filename):
                os.remove("old_" + short_i_b_filename)
                time.sleep(0.2)
            os.rename(short_i_b_filename, "old_" + short_i_b_filename)
            print_to_log("Renamed " + short_i_b_filename + " to " + "old_" + short_i_b_filename)
            time.sleep(0.2)
        os.rename(installed_base_report_filename, short_i_b_filename)
        print_to_log("Renamed " + installed_base_report_filename + " to " + short_i_b_filename)

    if os.path.isfile(service_contracts_report_filename):
        if os.path.isfile(short_s_c_filename):
            if os.path.isfile("old_" + short_s_c_filename):
                os.remove("old_" + short_s_c_filename)
                time.sleep(0.2)
            os.rename(short_s_c_filename, "old_" + short_s_c_filename)
            print_to_log("Renamed " + short_s_c_filename + " to " + "old_" + short_s_c_filename)
            time.sleep(0.2)
        os.rename(service_contracts_report_filename, short_s_c_filename)
        print_to_log("Renamed " + service_contracts_report_filename + " to " + short_s_c_filename)

    time.sleep(.5)


def check_response_profiles():
    try:
        with open("ranked_response_profiles.txt", encoding='utf-8') as f:
            lines = f.read().splitlines()
        HardwareContract.response_profiles_list = []
        for line in lines:
            if len(line.strip()) > 0:
                HardwareContract.response_profiles_list.append(line.strip())

    except FileNotFoundError:
        HardwareContract.response_profiles_list = [
            "2HR PREMIUM ONSITE",
            "4HR PREMIUM ONSITE",
            "4HR PARTS REPLACE",
            "4HR PARTS DELIVERY",
            "NBD PREMIUM ONSITE",
            "NBD PARTS REPLACE",
            "NBD PARTS DELIVERY",
            "QUARTERLY",
            "WARRANTY",
            "",
        ]
        print_to_log("Creating ranked_response_profiles.txt")
        with open("ranked_response_profiles.txt", 'w', encoding='utf-8') as f:
            for response_profile_item in HardwareContract.response_profiles_list:
                print(response_profile_item, file=f)

    print_to_log("Using ranked_response_profiles.txt: " + ', '.join(HardwareContract.response_profiles_list))

 #Betsy working on skip_entitlements_check for these products 5/3
    try:
        with open("products_to_skip_entitlement_check.txt", encoding='utf-8') as f:
            lines = f.read().splitlines()
            HardwareContract.products_to_exclude = []
            for line in lines:
                if line[0] != '#':
                    temp_words = line.strip().split(',', 2)
                    if len(line.strip()) > 0:
                        HardwareContract.products_to_exclude.append(line.strip())
                        print_to_log("Override entitlements for" + line.strip())
                    else:
                        print_to_log("Unable to override" + line.strip())
                else:
                    print_to_log("Unable to override entitlements for any product")

    except FileNotFoundError:
        HardwareContract.products_to_exclude = [
            "CL-LICENSEMANAGER"
        ]
        print_to_log("Creating products_to_skip_entitlement_check.txt")
        with open("products_to_skip_entitlement_check.txt", 'w', encoding='utf-8') as f:
            print("# products to exclude entitlements", file=f)
            print("# Note this will remove products (entitled or not)", file=f)
            print("CL-LICENSEMANAGER", file=f)
    print_to_log("Using products_to_skip_entitlement_check.txt: " + ', '.join(HardwareContract.products_to_exclude))


class IbDetails:

    def __init__(self):

        self.output_filename = ''

        self.ib_column_heading = {}
        self.sc_column_heading = {}
        self.column_width = {}
        self.system_by_serial = {}
        self.flash_by_serial = {}
        # self.switches_by_serial = {}
        self.system_by_hostname = {}
        self.unknown_hostname_systems = {}
        self.ib_hostnames = set()
        self.list_of_systems = []
        self.hostname_from_serial = {}
        self.companies_to_ignore = set()
        self.serials_to_ignore = set()
        self.hostnames_to_ignore = set()
        self.groups_to_ignore = set()
        self.sites_to_ignore = set()
        self.set_of_owners = set()
        self.set_of_sites = set()
        self.set_of_groups = set()
        self.last_week_owners = set()
        self.last_week_sites = set()
        self.last_week_groups = set()

        self.good_response_profiles = []

        self.check_groups = False
        self.customer_id_list = []
        global reseller_text

        check_input_files()
        check_response_profiles()
        self.map_hostnames()
        self.process_companies_to_ignore()
        self.process_serials_to_ignore()
        self.process_hostnames_to_ignore()
        self.process_groups_to_ignore()
        self.process_sites_to_ignore()
        self.process_ib_list()

        if os.path.isfile(sam_report_filename):
            use_sam_report = True
            site_text = "Installed At Site"
            cluster_serial_text = "Cluster Serial No"
            asup_date_text = "Last ASUP"
            product_family_text = "Product Family"
            platform_text = "Model"
            controller_eos_date_text = "Controller EOS"
            pvr_flag_text = "EOS PVR"
            pvr_date_text = "PVR End"
            first_eos_date_text = "First EOS"
            age_text = "System Age Years"
            ha_pair_flag_text = "HA Pair"
            company_text = "Serial Number Owner"
            print_to_log("Found sam.xlsx.")
            print_to_log("the following fields are being replaced with new headings")
            print_to_log(company_text)
            print_to_log(site_text)
            print_to_log(product_family_text)
            print_to_log(platform_text)
            print_to_log(ha_pair_flag_text)
            print_to_log(asup_date_text)
        else:
            use_sam_report = False
            print_to_log("sam.xlsx not found, will use legacy reports.")

        if use_sam_report:
            #try:
            self.ib_workbook = load_workbook(filename=sam_report_filename)
            assert "IB Products Detail" in self.ib_workbook.get_sheet_names(), "IB Products Detail missing from SAM report"
            #except:
                #print_to_log("Problem reading IB Products Detail, please try downloading a new sam.xlsx before running IB Scrub again")
                #return

            self.ib_worksheet = self.ib_workbook.get_sheet_by_name("IB Products Detail")

            if "IB Products Detail" in self.ib_worksheet.cell('A2').value and \
                    "Installed At Site" in self.ib_worksheet['B5'].value and \
                    self.ib_worksheet['E5'].value == "Serial Number" and \
                    self.ib_worksheet['AD5'].value == "# of Disks":
                print_to_log("Reading IB Products Detail tab from SAM report")
            else:
                print_to_log("Unexpected SAM report format, scrub cancelled")
                return

            my_row = 5
            my_column = 2

            while my_column <= last_data_column:
                self.ib_column_heading[self.ib_worksheet.cell(row=my_row, column=my_column).value] = my_column
                self.column_width[self.ib_worksheet.cell(row=my_row, column=my_column).value] = self.ib_worksheet.column_dimensions[get_column_letter(my_column)].width
                my_column += 1
            self.column_width[company_text] = self.column_width[site_text]

            # my_row = 9

            # save formatting
            self.title_font = self.ib_worksheet['A2'].font.copy()
            self.heading_font = self.ib_worksheet['B5'].font.copy()
            self.heading_alignment = self.ib_worksheet['B5'].alignment.copy()
            self.heading_fill = self.ib_worksheet['B5'].fill.copy()
            self.heading_border = self.ib_worksheet['B5'].border.copy()
            self.data_cell_font = self.ib_worksheet['B6'].font.copy()
            self.data_cell_alignment = self.ib_worksheet['B6'].alignment.copy()
            self.data_cell_fill = self.ib_worksheet['B6'].fill.copy()
            self.data_cell_border = self.ib_worksheet['B6'].border.copy()
        else:
            if os.path.isfile(short_i_b_filename):
                try:
                    self.ib_workbook = load_workbook(filename=short_i_b_filename)
                    assert data_sheet_name in self.ib_workbook.get_sheet_names(), data_sheet_name + " missing from Installed Base report"
                except:
                    print_to_log("Problem reading IB Details file, please try downloading a new version before running IB Scrub again")
                    return
            else:
                print_to_log("Missing IB report " + short_i_b_filename)
                return

            self.ib_worksheet = self.ib_workbook.get_sheet_by_name(data_sheet_name)

            if "Installed Base Details" in self.ib_worksheet.cell(ib_report_name_cell).value and \
                "3.0" in self.ib_worksheet.cell(ib_report_name_cell).value and \
                self.ib_worksheet['A8'].value == "Serial Number Owner Company" and \
                self.ib_worksheet['E8'].value == "Serial Number" and \
                    self.ib_worksheet['AD8'].value == "# of Disks":
                print_to_log("Reading Installed Base report v3.0")
            else:
                print_to_log("Unexpected Installed Base report format, scrub cancelled")
                return

            my_row = column_heading_row
            my_column = 1

            while my_column <= last_data_column:
                self.ib_column_heading[self.ib_worksheet.cell(row=my_row, column=my_column).value] = my_column
                self.column_width[self.ib_worksheet.cell(row=my_row, column=my_column).value] = self.ib_worksheet.column_dimensions[get_column_letter(my_column)].width
                my_column += 1

            my_row = 9

            # save formatting
            self.title_font = self.ib_worksheet['A7'].font.copy()
            self.heading_font = self.ib_worksheet['A8'].font.copy()
            self.heading_alignment = self.ib_worksheet['A8'].alignment.copy()
            self.heading_fill = self.ib_worksheet['A8'].fill.copy()
            self.heading_border = self.ib_worksheet['A8'].border.copy()
            self.data_cell_font = self.ib_worksheet['A9'].font.copy()
            self.data_cell_alignment = self.ib_worksheet['A9'].alignment.copy()
            self.data_cell_fill = self.ib_worksheet['A9'].fill.copy()
            self.data_cell_border = self.ib_worksheet['A9'].border.copy()

        if use_sam_report:

            #pandas, first try to see if sam and e-series exists, otherwise try just ontap
            try:
                df = pd.read_excel("sam.xlsx", sheetname="IB Products Detail", dtype='object')
                df.columns = df.iloc[3]
                df.columns = df.columns.fillna("Serial Number Owner")
                blank_sn = df["Serial Number"].isnull()
                blank_sites = df["Installed At Site"].isnull()
                temp = df[blank_sn & ~blank_sites]
                df["Serial Number Owner"] = temp["Installed At Site"]
                df["Serial Number Owner"] = df["Serial Number Owner"].fillna(method='ffill')
                header_row_pandas = df["Serial Number"] == "Serial Number"
                ontap_ib_final = df[~blank_sn & ~header_row_pandas]
                ontap_ib_final.fillna("", inplace=True)
                df = pd.read_excel("sam2.xlsx", sheetname="IB Products Detail", dtype='object')
                df.columns = df.iloc[3]
                df.columns = df.columns.fillna("Serial Number Owner")
                blank_sn = df["Serial Number"].isnull()
                blank_sites = df["Installed At Site"].isnull()
                temp = df[blank_sn & ~blank_sites]
                df["Serial Number Owner"] = temp["Installed At Site"]
                df["Serial Number Owner"] = df["Serial Number Owner"].fillna(method='ffill')
                header_row_pandas = df["Serial Number"] == "Serial Number"
                eseries_ib_final = df[~blank_sn & ~header_row_pandas]
                eseries_ib_final.fillna("", inplace=True)
                ib_final = ontap_ib_final.append(eseries_ib_final).drop_duplicates(keep='first', subset="Serial Number")
            except:
                df = pd.read_excel("sam.xlsx", sheetname="IB Products Detail", dtype='object')
                df.columns = df.iloc[3]
                df.columns = df.columns.fillna("Serial Number Owner")
                blank_sn = df["Serial Number"].isnull()
                blank_sites = df["Installed At Site"].isnull()
                temp = df[blank_sn & ~blank_sites]
                df["Serial Number Owner"] = temp["Installed At Site"]
                df["Serial Number Owner"] = df["Serial Number Owner"].fillna(method='ffill')
                header_row_pandas = df["Serial Number"] == "Serial Number"
                ib_final = df[~blank_sn & ~header_row_pandas]
                ib_final.fillna("", inplace=True)

            #read eseries

            #combine to create - ib_final

            header_row = list(ib_final.columns)
            data = []
            count = 0
            for index, row in ib_final.iterrows():
                count += 1
                if count % 1000 == 0:
                    print_to_log("Current row: " + str(count))
                record = {}
                for key in header_row:
                    record[key]=row[key]

                if record['Installed At Site'] and 'Installed At Site' in record['Installed At Site']:
                    continue

                if record['Serial Number'] is not None and len(record['Serial Number']) > 0 and record['Serial Number'] is not 'Serial Number':
                    # If we have a good Serial Number, go ahead and save the data
                    data.append(record)


            print_to_log("Done reading IB Products Detail")

        else:
            print_to_log("Reading " + short_i_b_filename)
            try:
                wb = load_workbook(short_i_b_filename, read_only=True)
            except:
                print_to_log("Problem reading IB Details file, please try downloading a new version before running IB Scrub again")
                return
            sheet = wb.active
            rows = sheet.rows
            header_row = [cell.value for cell in next(rows)]
            while header_row[0] != company_text:
                header_row = [cell.value for cell in next(rows)]
            data = []
            count = 0
            for row in rows:
                count += 1
                if count % 1000 == 0:
                    print_to_log("Current row: " + str(count))
                record = {}
                for key, cell in zip(header_row, row):
                    record[key] = cell.value
                data.append(record)
            print_to_log("Done reading ib.xlsx")

        self.serials_skipped = 0

        for row in data:

            if debug_it:
                print_to_log("Parsing row " + str(my_row) + " from IB report")
            # todo serial could be int

            try:
                serial = row[serial_text].strip()
            except AttributeError:
                serial = str(row[serial_text])

            if not serial:
                print_to_log("Blank serial on report row " + str(my_row))
                my_row += 1
                continue

            if serial in self.serials_to_ignore:
                if debug_it:
                    print_to_log("Ignore serial " + serial)
                my_row += 1
                self.serials_skipped += 1
                continue

            company = row[company_text]

            if company in self.companies_to_ignore:
                if debug_it:
                    print_to_log("Ignore system from " + company)
                my_row += 1
                self.serials_skipped += 1
                continue

            group = row[group_text]

            if group in self.groups_to_ignore:
                if debug_it:
                    print_to_log("Ignore system from " + group)
                my_row += 1
                self.serials_skipped += 1
                continue

            site = row[site_text]

            if site in self.sites_to_ignore:
                if debug_it:
                    print_to_log("Ignore system from " + site)
                my_row += 1
                self.serials_skipped += 1
                continue

            hostname = row[hostname_text].strip().lower()
            if hostname in self.hostnames_to_ignore:
                if debug_it:
                    print_to_log("Ignore hostname " + hostname)
                my_row += 1
                self.serials_skipped += 1
                continue

            if serial:
                system = NetAppSystem(serial, my_row)
                if serial in self.system_by_serial:
                    print_to_log("Warning: duplicate serial number in IB report will be ignored: " + serial)
                    my_row += 1
                    self.serials_skipped += 1
                    continue
                family = row[product_family_text]
                if family in ["FLASH CACHE", "CARD"]:
                    self.flash_by_serial[serial] = system
                    self.system_by_serial[serial] = system
                    self.list_of_systems.append(system)
                elif family in ["BROCADE"]:
                    # self.switches_by_serial[serial] = system
                    self.system_by_serial[serial] = system
                    self.list_of_systems.append(system)
                # elif family in ["FILER", "V-SERIES", "E-SERIES"]:
                elif True:
                    self.system_by_serial[serial] = system
                    self.list_of_systems.append(system)
                else:
                    print_to_log("Unexpected product for serial " + serial + ": " + family)
                    my_row += 1
                    self.serials_skipped += 1
                    continue

                system.hostname = hostname
                if system.serial in self.hostname_from_serial:
                    if system.hostname.strip().lower() != self.hostname_from_serial[system.serial].strip().lower():
                        print_to_log("Mapped " + serial + ":" + system.hostname.strip().lower() + " to " + self.hostname_from_serial[system.serial].strip().lower())
                    system.hostname = self.hostname_from_serial[system.serial].strip().lower()

                if family in ["FILER", "V-SERIES", "E-SERIES"]:
                    if system.hostname == "unknown":
                        self.unknown_hostname_systems[serial] = system
                    else:
                        if system.hostname in self.system_by_hostname:
                            print_to_log("Warning: hostname " + system.hostname + " is duplicated in IB, you must use serial number for notes, etc.")
                        self.system_by_hostname[system.hostname] = system

                system.set_owner(row[company_text])
                system.set_site(row[site_text])
                system.set_group(row[group_text])
                system.set_solution(row[solution_text])
                system.set_cluster_name(row[cluster_name_text])
                system.set_cluster_serial(row[cluster_serial_text])
                system.set_cluster_uuid(row[cluster_uuid_text])
                system.set_os_version(row[os_version_text])
                system.set_asup_status(row[asup_status_text])
                system.set_declined(row[declined_text])
                system.set_asup_date(row[asup_date_text])
                system.set_product_family(row[product_family_text])
                system.set_platform(row[platform_text])
                system.set_controller_eos_date(row[controller_eos_date_text])
                system.set_pvr_flag(row[pvr_flag_text])
                system.set_pvr_date(row[pvr_date_text])
                system.set_first_eos_date(row[first_eos_date_text])
                system.set_first_eos_part(row[first_eos_part_text])
                system.set_age(row[age_text])
                system.set_ha_pair_flag(row[ha_pair_flag_text])
                system.set_service_level(row[service_level_text])
                system.set_entitlement_status(row[entitlement_status_text])
                system.set_contact_name(row[contact_name_text])
                system.set_contact_number(row[contact_number_text])
                system.set_contact_email(row[contact_email_text])
                system.set_raw_tb(row[raw_tb_text])
                system.set_num_shelves(row[num_shelves_text])
                system.set_num_disks(row[num_disks_text])

                self.set_of_owners.add(system.owner)
                self.set_of_sites.add(system.site)
                self.set_of_groups.add(system.group)

            my_row += 1

        ib_report_row_count = my_row-9
        print_to_log("Processed " + str(ib_report_row_count) + " rows in IB Report")
        print_to_log("Processed " + str(len(self.system_by_serial)) + " serials from IB report")

        with open("serials-" + (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W") + ".txt", 'w', encoding='utf-8') as f:
            for serial in self.system_by_serial:
                print(serial, file=f)

        with open("sites-" + (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W") + ".txt", 'w', encoding='utf-8') as f:
            for site in self.set_of_sites:
                if site:
                    print(site, file=f)

        with open("owners-" + (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W") + ".txt", 'w', encoding='utf-8') as f:
            for owner in self.set_of_owners:
                if owner:
                    print(owner, file=f)

        with open("groups-" + (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W") + ".txt", 'w', encoding='utf-8') as f:
            for group in self.set_of_groups:
                if group:
                    print(group, file=f)

        if use_sam_report:
            try:
                assert "Service Contracts" in self.ib_workbook.get_sheet_names(), "Service Contracts missing from SAM report"
            except:
                print_to_log("Problem reading Service Contracts, please try downloading a new sam.xlsx before running IB Scrub again")
                return

            self.sc_worksheet = self.ib_workbook.get_sheet_by_name("Service Contracts")

            if "Service Contracts" in self.sc_worksheet.cell("A2").value and \
                "Serial Number Owner" in self.sc_worksheet['B4'].value and \
                self.sc_worksheet['G4'].value == "Serial Number" and \
                    (self.sc_worksheet['R4'].value == "HW Service Level Status" or self.sc_worksheet['R4'].value == "Support Edge Offering Status"):
                print_to_log("Reading Service Contracts tab from SAM report")
                site_text = "Installed At Site Name"
                print_to_log("the following is being replaced with new headings")
                print_to_log(site_text)

            else:
                print_to_log("Unexpected Service Contracts, scrub cancelled")
                return
            hw_service_level_status_text = self.sc_worksheet['R4'].value
            my_row = 4  # column headings
            my_column = 1

            while my_column <= sc_last_data_column + 1:  # sam.xlsx has blank column A
                if self.sc_worksheet.cell(row=my_row, column=my_column).value:
                    self.sc_column_heading[self.sc_worksheet.cell(row=my_row, column=my_column).value] = my_column
                    self.column_width[self.sc_worksheet.cell(row=my_row, column=my_column).value] = self.sc_worksheet.column_dimensions[get_column_letter(my_column)].width
                    self.column_width["NRD " + self.sc_worksheet.cell(row=my_row, column=my_column).value] = self.sc_worksheet.column_dimensions[get_column_letter(my_column)].width
                my_column += 1

            my_row = 9
        else:
            if os.path.isfile(short_s_c_filename):
                try:
                    self.sc_workbook = load_workbook(filename=short_s_c_filename)
                    assert data_sheet_name in self.sc_workbook.get_sheet_names(), data_sheet_name + " missing from Service Contracts report"
                except:
                    print_to_log("Problem reading Service Contracts file, please try downloading a new version before running IB Scrub again")
                    return
            else:
                print_to_log("Missing SC report " + short_s_c_filename)
                return

            self.sc_worksheet = self.sc_workbook.get_sheet_by_name(data_sheet_name)

            if "Service Contracts" in self.sc_worksheet.cell(sc_report_name_cell).value and \
                self.sc_worksheet['A8'].value == "Serial Number Owner Company" and \
                self.sc_worksheet['F8'].value == "Serial Number" and \
                    (self.sc_worksheet['S8'].value == "HW Service Level Status" or self.sc_worksheet['S8'].value == "Support Edge Offering Status"):
                print_to_log("Reading Service Contracts report")
            else:
                print_to_log("Unexpected Service Contracts, scrub cancelled")
                return
            hw_service_level_status_text = self.sc_worksheet['S8'].value
            my_row = column_heading_row
            my_column = 1

            while my_column <= sc_last_data_column:
                self.sc_column_heading[self.sc_worksheet.cell(row=my_row, column=my_column).value] = my_column
                self.column_width[self.sc_worksheet.cell(row=my_row, column=my_column).value] = self.sc_worksheet.column_dimensions[get_column_letter(my_column)].width
                self.column_width["NRD " + self.sc_worksheet.cell(row=my_row, column=my_column).value] = self.sc_worksheet.column_dimensions[get_column_letter(my_column)].width
                my_column += 1

            if "Months till Expire" in self.column_width:
                self.column_width["Months Till Expire"] = self.column_width["Months till Expire"]
                self.column_width["NRD Months Till Expire"] = self.column_width["Months till Expire"]

            my_row = 9

        if use_sam_report:
            #use pandas
            try:
                sc_ontap = pd.read_excel("sam.xlsx", sheetname="Service Contracts", header=2, dtype='object')
                sc_ontap["Contract Status"] = sc_ontap["Months till Expire"]
                sc_ontap["Contract Extended Date"] = sc_ontap["Contract End Date"]
                sc_ontap["Serial Number"] = sc_ontap["Serial Number"].astype("object")
                sc_ontap.fillna("", inplace=True)
                sc_ontap = sc_ontap.drop(sc_ontap.columns[0], axis=1)
                sc_eseries = pd.read_excel("sam2.xlsx", sheetname="Service Contracts", header=2, dtype='object')
                sc_eseries["Contract Status"] = sc_eseries["Months till Expire"]
                sc_eseries["Contract Extended Date"] = sc_eseries["Contract End Date"]
                sc_eseries["Serial Number"] = sc_eseries["Serial Number"].astype("object")
                sc_eseries.fillna("", inplace=True)
                sc_eseries = sc_eseries.drop(sc_eseries.columns[0], axis=1)
                sc = sc_ontap.append(sc_eseries).drop_duplicates(keep='first')
            except:
                sc = pd.read_excel("sam.xlsx", sheetname="Service Contracts", header=2, dtype='object')
                sc["Contract Status"] = sc["Months till Expire"]
                sc["Contract Extended Date"] = sc["Contract End Date"]
                sc["Serial Number"] = sc["Serial Number"].astype("object")
                sc.fillna("", inplace=True)

            header_row = list(sc.columns)
            data = []
            count = 0
            for index, row in sc.iterrows():
                count += 1
                if count % 1000 == 0:
                    print_to_log("Current row: " + str(count))
                record = {}
                for key in header_row:
                    record[key]=row[key]
                if record[serial_text]:
                    if "Months till Expire" in record:
                        record["Months Till Expire"] = record["Months till Expire"]
                    data.append(record)
        else:
            print_to_log("Reading " + short_s_c_filename)
            try:
                wb = load_workbook(short_s_c_filename, read_only=True)
            except:
                print_to_log("Problem reading Service Contracts file, please try downloading a new version before running IB Scrub again")
                return
            sheet = wb.active
            rows = sheet.rows
            header_row = [cell.value for cell in next(rows)]
            while header_row[0] != company_text:
                header_row = [cell.value for cell in next(rows)]
            data = []
            count = 0
            for row in rows:
                count += 1
                if count % 1000 == 0:
                    print_to_log("Current row: " + str(count))
                record = {}
                for key, cell in zip(header_row, row):
                    record[key] = cell.value
                data.append(record)

        for row in data:
            if debug_it:
                print_to_log("Parsing row " + str(my_row) + " from SC report")

            company = row[company_text]

            if company in self.companies_to_ignore:
                if debug_it:
                    print_to_log("Ignore system from " + company)
                my_row += 1
                continue

            group = row[group_text]

            if group in self.groups_to_ignore:
                if debug_it:
                    print_to_log("Ignore system from " + group)
                my_row += 1
                continue

            site = row[site_text]

            if site in self.sites_to_ignore:
                if debug_it:
                    print_to_log("Ignore system from " + site)
                my_row += 1
                continue

            hostname = row[hostname_text].strip().lower()
            if hostname in self.hostnames_to_ignore:
                if debug_it:
                    print_to_log("Ignore hostname " + hostname)
                my_row += 1
                continue

            service_type = row[service_type_text]

            if service_type and service_type in ["SW"]:
                if debug_it:
                    print_to_log("Ignore service type " + service_type)
                my_row += 1
                continue

            serial = str(row[serial_text]).strip()

            if not serial:
                print_to_log("Blank serial on SC report row " + str(my_row))
                my_row += 1
                continue

            if serial in self.serials_to_ignore:
                if debug_it:
                    print_to_log("Ignore serial " + serial)
                my_row += 1
                continue

            if debug_it:
                print_to_log("sc_report - Found serial " + serial)

            # if serial in self.system_by_serial and self.system_by_serial[serial].product_family != "FLASH CACHE":
            # treat flash cache and all products the same as controllers
            if serial in self.system_by_serial:
                # todo check for longest duration contract
                if service_type == "NRD":
                    system = self.system_by_serial[serial]
                    if system.nrd_service_contract_id:
                        if debug_it:
                            print_to_log("Another NRD contracts for serial " + system.serial)
                        try:
                            if system.nrd_months_till_expire < row[months_till_expire_text]:
                                print_to_log("Replacing NRD contract " + str(system.nrd_months_till_expire) + " with " + str(row[months_till_expire_text]))
                                system.set_nrd_service_contract_id(row[service_contract_id_text])
                                system.set_nrd_contract_end_date(row[contract_end_date_text])
                                system.set_nrd_months_till_expire(row[months_till_expire_text])
                                system.set_nrd_hw_service_level_status(row[hw_service_level_status_text])
                        except:
                            print_to_log("Problem comparing NRD contracts for serial " + system.serial)
                    else:
                        if debug_it:
                            print_to_log("New NRD contracts for serial " + system.serial)
                        system.set_nrd_service_contract_id(row[service_contract_id_text])
                        system.set_nrd_contract_end_date(row[contract_end_date_text])
                        system.set_nrd_months_till_expire(row[months_till_expire_text])
                        system.set_nrd_hw_service_level_status(row[hw_service_level_status_text])
                elif service_type == "HW" or service_type == '' or service_type is None:
                    if debug_it:
                        print_to_log("row " + str(my_row))
                    system = self.system_by_serial[serial]
                    my_service_contract_id = row[service_contract_id_text]
                    my_warranty_end_date = row[warranty_end_date_text]
                    my_contract_end_date = row[contract_end_date_text]
                    new_contract = HardwareContract(row[service_level_text],
                                                    my_warranty_end_date,
                                                    "HW",
                                                    row[response_profile_text],
                                                    my_service_contract_id,
                                                    my_contract_end_date,
                                                    row[months_till_expire_text],  # check
                                                    #row[contract_status_text], # removed
                                                    row[hw_service_level_status_text],
                                                    row[entitlement_status_text],
                                                    serial)
                    # print_to_log("New contract: " + str(new_contract))
                    previous_contract = HardwareContract(system.service_level,
                                                         system.warranty_end_date,
                                                         system.service_type,
                                                         system.response_profile,
                                                         system.service_contract_id,
                                                         system.contract_end_date,
                                                         system.months_till_expire,
                                                         #removed from SAM - system.contract_status,
                                                         system.hw_service_level_status,
                                                         system.sc_entitlement_status,
                                                         serial)
                    # print_to_log("Previous contract: " + str(previous_contract))
                    previous_longest_contract = HardwareContract(system.longest_service_level,
                                                                 system.longest_warranty_end_date,
                                                                 system.longest_service_type,
                                                                 system.longest_response_profile,
                                                                 system.longest_service_contract_id,
                                                                 system.longest_contract_end_date,
                                                                 system.longest_months_till_expire,
                                                                 #removed from SAM - system.longest_contract_status,
                                                                 system.longest_hw_service_level_status,
                                                                 system.longest_sc_entitlement_status,
                                                                 serial)

                    # print_to_log("Calculating effective contract for " + system.serial)
                    effective_contract = new_contract & previous_contract

                    if previous_contract and effective_contract != previous_contract:
                        print_to_log("Active Contract " + str(previous_contract) + " replaced by " + str(effective_contract) + " for serial " + system.serial)
                    # print_to_log("Effective contract " + str(effective_contract))

                    system.set_service_level(effective_contract.service_level)
                    system.set_warranty_end_date(effective_contract.warranty_end_date)
                    system.set_service_type(effective_contract.service_type)
                    system.set_response_profile(effective_contract.response_profile)
                    system.set_service_contract_id(effective_contract.service_contract_id)
                    system.set_contract_end_date(effective_contract.contract_end_date)
                    system.set_months_till_expire(effective_contract.months_till_expire)
                    # removed from SAM -system.set_contract_status(effective_contract.contract_status)
                    system.set_hw_service_level_status(effective_contract.hw_service_level_status)
                    system.set_sc_entitlement_status(effective_contract.sc_entitlement_status)

                    # print_to_log("checking longer contract for serial " + system.serial)
                    longer_contract = new_contract | previous_longest_contract
                    if longer_contract is None:
                        longer_contract = effective_contract
                        print_to_log("Warning: could not determine longest contract using effective contract")
                    if previous_longest_contract and longer_contract != previous_longest_contract:
                        print_to_log("Longest Contract " + str(previous_longest_contract) + " replaced by " + str(longer_contract) + " for serial " + system.serial)
                    # print_to_log("Effective contract " + str(effective_contract))

                    system.set_longest_service_level(longer_contract.service_level)
                    system.set_longest_warranty_end_date(longer_contract.warranty_end_date)
                    system.set_longest_service_type(longer_contract.service_type)
                    system.set_longest_response_profile(longer_contract.response_profile)
                    system.set_longest_service_contract_id(longer_contract.service_contract_id)
                    system.set_longest_contract_end_date(longer_contract.contract_end_date)
                    system.set_longest_months_till_expire(longer_contract.months_till_expire)
                    #removed from SAM -system.set_longest_contract_status(longer_contract.contract_status)
                    system.set_longest_hw_service_level_status(longer_contract.hw_service_level_status)
                    system.set_longest_sc_entitlement_status(longer_contract.sc_entitlement_status)

                    # system.set_service_level(csv_row[service_level_text])
                    # system.set_warranty_end_date(csv_row[warranty_end_date_text])
                    # system.set_service_type(csv_row[service_type_text])
                    # system.set_response_profile(csv_row[response_profile_text])
                    # system.set_service_contract_id(csv_row[service_contract_id_text])
                    # system.set_contract_end_date(csv_row[contract_end_date_text])
                    # system.set_months_till_expire(csv_row[months_till_expire_text])
                    # system.set_contract_status(csv_row[contract_status_text])
                    # system.set_hw_service_level_status(csv_row[hw_service_level_status_text])
                    # system.set_sc_entitlement_status(csv_row[entitlement_status_text])
                else:
                    if service_type:
                        print_to_log("Unexpected service type: " + service_type + " on row " + str(my_row))
                    else:
                        print_to_log("Unexpected service type on row " + str(my_row))
                        print_to_log(str(type(service_type)))
            elif serial in self.flash_by_serial:
                # don't think this is used anymore as flash is treated as a system
                if service_type == "NRD":
                    system = self.flash_by_serial[serial]
                    system.set_nrd_service_contract_id(row[service_contract_id_text])
                    system.set_nrd_contract_end_date(row[contract_end_date_text])
                    system.set_nrd_months_till_expire(row[months_till_expire_text])
                    system.set_nrd_hw_service_level_status(row[hw_service_level_status_text])
                else:
                    # todo check for unexpected service_type
                    system = self.flash_by_serial[serial]
                    system.set_service_level(row[service_level_text])
                    system.set_warranty_end_date(row[warranty_end_date_text])
                    system.set_service_type(row[service_type_text])
                    system.set_response_profile(row[response_profile_text])
                    system.set_service_contract_id(row[service_contract_id_text])
                    system.set_contract_end_date(row[contract_end_date_text])
                    system.set_months_till_expire(row[months_till_expire_text])
                    #removed from SAM -system.set_contract_status(row[contract_status_text])
                    system.set_hw_service_level_status(row[hw_service_level_status_text])
                    system.set_sc_entitlement_status(row[entitlement_status_text])

            else:
                print_to_log("Unexpected serial in SC Report: " + serial)
            my_row += 1

        sc_report_row_count = my_row-1
        print_to_log("Processed " + str(sc_report_row_count) + " rows in SC Report")

        if os.path.isfile(score_report_filesname):
            try:
                self.score_workbook = load_workbook(filename=score_report_filesname)
            except:
                print_to_log("Problem reading SCORE file, please try downloading a new version before running IB Scrub again")
                return

            self.score_worksheet = self.score_workbook.active

            if self.score_worksheet['A1'].value == "Sales Geography" and \
                self.score_worksheet['V1'].value == "Serial Number" and \
                    self.score_worksheet['W1'].value == "Partner Serial Number":
                print_to_log("Reading SCORE report")

                my_row = 1
                my_column = 1
                score_column_heading = {}
                score_column_width = {}
                while self.score_worksheet.cell(row=my_row, column=my_column).value:
                    score_column_heading[self.score_worksheet.cell(row=my_row, column=my_column).value] = my_column
                    # print_to_log(self.score_worksheet.cell(row=my_row, column=my_column).value + " is column " + str(my_column))
                    score_column_width[self.score_worksheet.cell(row=my_row, column=my_column).value] = self.score_worksheet.column_dimensions[get_column_letter(my_column)].width
                    my_column += 1

                if reseller_text not in score_column_heading:
                    if "Reseller As Sold Company" in score_column_heading:
                        reseller_text = "Reseller As Sold Company"
                    else:
                        print_to_log("Unexpected SCORE format, please contact Neil Maldonado")
                        raise ValueError
                my_row += 1

                for column_heading in score_column_width:
                    if column_heading not in self.column_width:
                        self.column_width[column_heading] = score_column_width[column_heading]
                self.column_width["Partner Serial Number"] = self.column_width["Serial Number"]
                self.column_width["Old Serial Number"] = self.column_width["Serial Number"]
                self.column_width["Original SO#"] = self.column_width["Serial Number"]
                self.column_width["Filer Use"] = self.column_width["Group Name"]

                print_to_log("Reading " + score_report_filesname)
                try:
                    wb = load_workbook(score_report_filesname, read_only=True)
                except:
                    print_to_log("Problem reading Service Contracts file, please try downloading a new version before running IB Scrub again")
                    return
                sheet = wb.active
                rows = sheet.rows
                header_row = [cell.value for cell in next(rows)]
                while header_row[0] != "Sales Geography":
                    header_row = [cell.value for cell in next(rows)]
                data = []
                count = 0
                for row in rows:
                    count += 1
                    if count % 1000 == 0:
                        print_to_log("Current row: " + str(count))
                    record = {}
                    for key, cell in zip(header_row, row):
                        record[key] = cell.value
                    data.append(record)

                for row in data:
                    if debug_it:
                        print_to_log("Parsing row " + str(my_row) + " from SCORE report")
                    if not row[serial_text]:
                        continue
                    try:
                        serial = str(int(row[serial_text]))
                    except ValueError:
                        serial = row[serial_text]

                    if not serial:
                        print_to_log("Blank serial on report row " + str(my_row))
                        my_row += 1
                        continue

                    if serial in self.serials_to_ignore:
                        if debug_it:
                            print_to_log("Ignore serial " + serial)
                        my_row += 1
                        continue

                    if serial in self.system_by_serial:
                        system = self.system_by_serial[serial]
                        system.set_partner_serial(row["Partner Serial Number"])
                        system.set_account_name(row[account_name_text])
                        # print_to_log(system.account_name)
                        system.set_end_customer(row[end_customer_text])
                        # print_to_log(system.end_customer)
                        system.set_customer_geo(row[customer_geo_text])
                        # print_to_log(system.customer_geo)
                        system.set_customer_country(row[customer_country_text])
                        # print_to_log(system.customer_country)
                        system.set_customer_city(row[customer_city_text])
                        # print_to_log(system.customer_city)
                        system.set_original_so(row[original_so_text].replace(""""@"';">""", ""))
                        # print_to_log(system.original_so)
                        system.set_shipped_date(row[shipped_date_text])
                        # print_to_log(str(system.shipped_date))
                        system.set_software_service_end_date(row[software_service_end_date_text])
                        # print_to_log(str(system.software_service_end_date))
                        system.set_reseller(row[reseller_text])
                        # print_to_log(system.reseller)
                        system.set_solidfire_tag(row[solidfire_tag_text])
                        # print_to_log(system.solidfire_tag_text)
                    my_row += 1
            else:
                print_to_log("Unexpected SCORE report format, skipping")

        else:
            print_to_log("score.xlsx not available")
            print_to_log("Use eBI Installed Base Analytics Dashboard, Service Contract Renewal, Installed Base Report")
            print_to_log("(file name will be SCORE  - Install Base Report QE.xls, open and save as score.xlsx)")

        try:
            with open(ib_notes_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()
            for line in lines:
                if len(line.strip()) > 0:
                    if line[0] != '#':
                        temp_words = line.strip().split(',', 1)
                        if len(temp_words) == 2:
                            if temp_words[0].strip().lower() in self.system_by_hostname:
                                self.system_by_hostname[temp_words[0].strip().lower()].add_note(temp_words[1])
                                if debug_it:
                                    print_to_log("Added note " + temp_words[0].strip().lower() + ": " + temp_words[1])
                            elif temp_words[0].strip() in self.system_by_serial:
                                self.system_by_serial[temp_words[0].strip()].add_note(temp_words[1])
                                if debug_it:
                                    print_to_log("Added note " + temp_words[0].strip() + ": " + temp_words[1])
                            elif temp_words[0].strip().lower() in self.system_by_serial:
                                self.system_by_serial[temp_words[0].strip().lower()].add_note(temp_words[1])
                                if debug_it:
                                    print_to_log("Added note " + temp_words[0].strip().lower() + ": " + temp_words[1])
                            else:
                                print_to_log("System not found in ib.txt: " + line.strip())

        except FileNotFoundError:
            with open(ib_notes_filename, 'w', encoding='utf-8') as f:
                print("#use the following format (without the #)", file=f)
                print("#serial,notes for this serial", file=f)
                print("#hostname,notes for this hostname", file=f)
                print("#comments denoted by line starting with '#'", file=f)
                print_to_log("Created " + ib_notes_filename)

        # create output file
        self.output_filename = "ib_scrub-" + time.strftime("%Y%m%d%H%M%S") + ".xlsx"
        wb = Workbook()
        ws = wb.get_sheet_by_name('Sheet')
        ws.title = "IB Scrub"

        ws['A1'] = "IB Team Installed Base Scrub v1.9"
        # ws['A1'].font = Font(name='Arial', b=True, color=Color(rgb='FF0066CC'), sz=12.0)
        ws['A1'].font = self.title_font

        # SAM report

        # columns to exclude
        columns_to_exclude = []
        try:
            with open("columns_to_exclude.txt", encoding='utf-8') as f:
                lines = f.read().splitlines()
            for line in lines:
                columns_to_exclude.append(line.strip())

        except FileNotFoundError:
            with open("columns_to_exclude.txt", 'w', encoding='utf-8') as f:
                print("columns that appear in this file will be excluded from reports, spelling and case sensitive", file=f)
                print(solution_text, file=f)
                columns_to_exclude.append(solution_text)
                print(cluster_name_text, file=f)
                columns_to_exclude.append(cluster_name_text)
                print(cluster_uuid_text, file=f)
                columns_to_exclude.append(cluster_uuid_text)
                print(first_eos_date_text, file=f)
                columns_to_exclude.append(first_eos_date_text)
                print(first_eos_part_text, file=f)
                columns_to_exclude.append(first_eos_part_text)
                print(declined_text, file=f)
                columns_to_exclude.append(declined_text)
                print(raw_tb_text, file=f)
                columns_to_exclude.append(raw_tb_text)

        column_list = [
            (serial_text, "serial", "0"),
            ("Old Serial Number", "old_serial", "0"),
            (account_name_text, "account_name", "General"),
            (company_text, "owner", "General"),
            (end_customer_text, "end_customer", "General"),
            (site_text, "site", "General"),
            (customer_geo_text, "customer_geo", "General"),
            (customer_country_text, "customer_country", "General"),
            (customer_city_text, "customer_city", "General"),
            (group_text, "group", "General"),
            (solution_text, "solution", "General"),
            (hostname_text, "hostname", "General"),
            ("Partner Serial Number", "partner_serial", "0"),  # partner serial number from SCORE report
            (cluster_name_text, "cluster_name", "General"),
            (cluster_serial_text, "cluster_serial", "General"),
            (cluster_uuid_text, "cluster_uuid", "General"),
            (os_version_text, "os_version", "General"),
            (asup_status_text, "asup_status", "General"),
            (declined_text, "declined", "General"),
            (asup_date_text, "asup_date", "mmm dd, yyyy"),
            (product_family_text, "product_family", "General"),
            (platform_text, "platform", "General"),
            (controller_eos_date_text, "controller_eos_date", "mmm dd, yyyy"),
            (pvr_flag_text, "pvr_flag", "General"),
            (pvr_date_text, "pvr_date", "mmm dd, yyyy"),
            (first_eos_date_text, "first_eos_date", "mmm dd, yyyy"),
            (first_eos_part_text, "first_eos_part", "General"),
            (age_text, "age", "0.00"),
            (shipped_date_text, "shipped_date", "mmm dd, yyyy"),
            ("Original SO#", "original_so", "0"),
            (ha_pair_flag_text, "ha_pair_flag", "General"),
            (service_level_text, "service_level", "General"),
            (entitlement_status_text, "entitlement_status", "General"),
            (contact_name_text, "contact_name", "General"),
            (contact_number_text, "contact_number", "General"),
            (contact_email_text, "contact_email", "General"),
            (raw_tb_text, "raw_tb", "0"),
            (num_shelves_text, "num_shelves", "0"),
            (num_disks_text, "num_disks", "0"),
            (nrd_service_contract_id_text, "nrd_service_contract_id", "General"),
            (nrd_contract_end_date_text, "nrd_contract_end_date", "mmm dd, yyyy"),
            (nrd_months_till_expire_text, "nrd_months_till_expire", "0"),
            (nrd_hw_service_level_status_text, "nrd_hw_service_level_status", "General"),
            (warranty_end_date_text, "warranty_end_date", "mmm dd, yyyy"),
            (service_type_text, "service_type", "General"),
            (response_profile_text, "response_profile", "General"),
            (service_contract_id_text, "service_contract_id", "General"),
            (contract_end_date_text, "contract_end_date", "mmm dd, yyyy"),
            (months_till_expire_text, "months_till_expire", "0"),
            # (contract_status_text, "contract_status", "General"),
            (hw_service_level_status_text, "hw_service_level_status", "General"),
            (software_service_end_date_text, "software_service_end_date", "mmm dd, yyyy"),
            (reseller_text, "reseller", "General"),
            ("Filer Use", "filer_use", "0"),
            ("Longest " + response_profile_text, "longest_response_profile", "General"),
            ("Longest " + service_contract_id_text, "longest_service_contract_id", "General"),
            ("Longest " + contract_end_date_text, "longest_contract_end_date", "mmm dd, yyyy"),
            ("Longest " + months_till_expire_text, "longest_months_till_expire", "0"),
            (solidfire_tag_text, "solidfire_tag", "General") #From SCORE Report
        ]

        attribute_text = {}
        attribute_format_string = {}
        for text, attribute_name, format_string in column_list:
            attribute_text[attribute_name] = text
            attribute_format_string[attribute_name] = format_string

        try:
            with open(override_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()
            for line in lines:
                if line[0] != '#':
                    temp_words = line.strip().split(',', 2)
                    if len(temp_words) == 3:
                        host = temp_words[0].strip().lower()
                        serial = temp_words[0].strip()
                        attr_name = temp_words[1].strip().lower()
                        value = temp_words[2].strip()
                        if host in self.system_by_hostname and hasattr(self.system_by_hostname[host], attr_name):
                            print_to_log("Override  " + host + " " + attr_name + " " + getattr(self.system_by_hostname[host], attr_name) + " with " + value)
                            setattr(self.system_by_hostname[host], attr_name, value)
                        elif serial in self.system_by_serial and hasattr(self.system_by_serial[serial], attr_name):
                            print_to_log("Override  " + serial + " " + attr_name + " " + getattr(self.system_by_serial[serial], attr_name) + " with " + value)
                            setattr(self.system_by_serial[serial], attr_name, value)
                        elif serial.lower() in self.system_by_serial and hasattr(self.system_by_serial[serial.lower()], attr_name):
                            print_to_log("Override  " + serial.lower() + " " + attr_name + " " + getattr(self.system_by_serial[serial.lower()], attr_name) + " with " + value)
                            setattr(self.system_by_serial[serial.lower()], attr_name, value)
                        else:
                            print_to_log("Unable to override: " + attr_name + " on " + host + "/" + serial)
                    else:
                        print_to_log("Unexpected override: " + line.strip())
        except FileNotFoundError:
            with open(override_filename, 'w', encoding='utf-8') as f:
                print("#override format:", file=f)
                print("#serial,attribute,value", file=f)
                print("# or", file=f)
                print("#hostname,attribute,value", file=f)
                print("#valid attributes:", file=f)
                for column_heading, attr_name, format_string in column_list:
                    print("#" + attr_name, file=f)
                print("#remove the '#' on your overrides please :)", file=f)
                print_to_log("Created " + override_filename)

        my_column = 1
        my_row = 2
        for column_heading, attr_name, format_string in column_list:
            if column_heading not in self.column_width:
                self.column_width[column_heading] = self.column_width["Serial Number"]
            if column_heading in columns_to_exclude:
                print_to_log("Excluding column " + column_heading)
                continue
                #Betsy overrides column headings
            if column_heading == "Contract End Date":
                cell = ws.cell(row=my_row, column=my_column, value="Best Service Level End Date")
            elif column_heading == "Response Profile":
                cell = ws.cell(row=my_row, column=my_column, value="Best Response Profile")
            elif column_heading == "Contract Extended Date":
                cell = ws.cell(row=my_row, column=my_column, value="Best Contract Extend Date")
            elif column_heading == "Months Till Expire":
                cell = ws.cell(row=my_row, column=my_column, value="Best Months Till Expire")
            elif column_heading == "Service Contract ID":
                cell = ws.cell(row=my_row, column=my_column, value="Best Service Contract ID")
            elif column_heading == "Cluster Serial No":
                cell = ws.cell(row=my_row, column=my_column, value="Cluster Serial Number")
            elif column_heading == "Last ASUP":
                cell = ws.cell(row=my_row, column=my_column, value="ASUP Last Generate Date")
            elif column_heading == "Product Family":
                cell = ws.cell(row=my_row, column=my_column, value="Installed At Product Family")
            elif column_heading == "Model":
                cell = ws.cell(row=my_row, column=my_column, value="Installed At Platform")
            elif column_heading == "Controller EOS":
                cell = ws.cell(row=my_row, column=my_column, value="Controller EOS Date")
            elif column_heading == "EOS PVR":
                cell = ws.cell(row=my_row, column=my_column, value="EOS PVR Flag")
            elif column_heading == "PVR End":
                cell = ws.cell(row=my_row, column=my_column, value="PVR End Date")
            elif column_heading == "First EOS":
                cell = ws.cell(row=my_row, column=my_column, value="First EOS Date")
            elif column_heading == "System Age Years":
                cell = ws.cell(row=my_row, column=my_column, value="System Age in Years")
            elif column_heading == "HA Pair":
                cell = ws.cell(row=my_row, column=my_column, value="HA Pair Flag")
            elif column_heading == "Serial Number Owner":
                cell = ws.cell(row=my_row, column=my_column, value="Serial Number Owner Company")
            else:
                cell = ws.cell(row=my_row, column=my_column, value=column_heading)

            cell.font = self.heading_font
            cell.alignment = self.heading_alignment
            cell.border = self.heading_border
            cell.fill = self.heading_fill
            ws.column_dimensions[get_column_letter(my_column)].width = self.column_width[column_heading]
            if attr_name == "controller_eos_date":
                my_column += 1
                cell = ws.cell(row=my_row, column=my_column, value="Months Till Controller EOS")
                cell.font = self.heading_font
                cell.alignment = self.heading_alignment
                cell.border = self.heading_border
                cell.fill = self.heading_fill
                ws.column_dimensions[get_column_letter(my_column)].width = self.column_width[first_eos_part_text]

            if attr_name == "pvr_date":
                my_column += 1
                cell = ws.cell(row=my_row, column=my_column, value="PVR Months Left")
                cell.font = self.heading_font
                cell.alignment = self.heading_alignment
                cell.border = self.heading_border
                cell.fill = self.heading_fill
                ws.column_dimensions[get_column_letter(my_column)].width = self.column_width[first_eos_part_text]

            if attr_name == "first_eos_part":
                my_column += 1
                cell = ws.cell(row=my_row, column=my_column, value="Months Till First EOS")
                cell.font = self.heading_font
                cell.alignment = self.heading_alignment
                cell.border = self.heading_border
                cell.fill = self.heading_fill
                ws.column_dimensions[get_column_letter(my_column)].width = self.column_width[first_eos_part_text]

            if attr_name == "partner_serial":
                my_column += 1
                cell = ws.cell(row=my_row, column=my_column, value="Partner Hostname")
                cell.font = self.heading_font
                cell.alignment = self.heading_alignment
                cell.border = self.heading_border
                cell.fill = self.heading_fill
                ws.column_dimensions[get_column_letter(my_column)].width = self.column_width[hostname_text]

            my_column += 1
        cell = ws.cell(row=my_row, column=my_column, value="Notes")
        cell.font = self.heading_font
        cell.alignment = self.heading_alignment
        cell.border = self.heading_border
        cell.fill = self.heading_fill
        ws.column_dimensions[get_column_letter(my_column)].width = 50

        my_row = 3
        expired_count = 0
        owner_companies = set()
        sites = set()
        systems_missing_group = []
        asup_off_count = 0
        eos_pvr_count = 0
        response_profile_counts = {}
        months_till_expire_count = {}
        months_till_expire_count["Less than 12mos:"] = 0
        months_till_expire_count["Less than 6mos:"] = 0
        months_till_expire_count["Less than 3mos:"] = 0
        nrd_contracts_count = 0
        expired_nrd_count = 0
        nrd_months_till_expire_count = {}
        nrd_months_till_expire_count["Less than 12mos:"] = 0
        nrd_months_till_expire_count["Less than 6mos:"] = 0
        nrd_months_till_expire_count["Less than 3mos:"] = 0
        months_until_eos_count = {}
        months_until_eos_count["Controllers:"] = 0
        months_until_eos_count["Controllers EOS next 12mos:"] = 0
        months_until_eos_count["Controllers EOS 13-24mos:"] = 0
        months_until_eos_count["Other:"] = 0
        months_until_eos_count["Other EOS next 12mos:"] = 0
        months_until_eos_count["Other EOS 13-24mos:"] = 0
        eos_keys = ["Controllers:",
                    "Controllers EOS next 12mos:",
                    "Controllers EOS 13-24mos:",
                    "Other:",
                    "Other EOS next 12mos:",
                    "Other EOS 13-24mos:"]

        print_to_log("Merging reports...")

        # Betsy adding the count of serials at each site, group, and owner 5.3
        count_by_site = {}
        count_by_group = {}
        count_by_owner = {}
        expired_list = []
        bad_response_profile_list = []
        count = 0
        for system in self.list_of_systems:
            count += 1
            if count % 1000 == 0:
                print_to_log("Finished " + str(count) + " systems")
            my_column = 1
            if "expired" in system.hw_service_level_status.lower() or ("check" in system.hw_service_level_status.lower() and system.months_till_expire <= 0):
                expired_count += 1
                expired_list.append(system)

            owner_companies.add(system.owner)
            sites.add(system.site)
            if system.hostname.lower() != "unknown" and not system.group:
                systems_missing_group.append(system)
        # Betsy this counts the number of serials for each group
            if system.group not in count_by_group:
                count_by_group[system.group] = 0
            count_by_group[system.group] += 1
        # Betsy This counts the number of serials under each owner
            if system.owner not in count_by_owner:
                count_by_owner[system.owner] = 0
            count_by_owner[system.owner] += 1
        # Betsy - This counts the number of serials at each site
            if system.site not in count_by_site:
               count_by_site[system.site] = 0
            count_by_site[system.site] += 1

            # if 'off' in system.asup_status.lower() and system.product_family not in ["BROCADE", "FLASH CACHE"]:
            if 'off' in system.asup_status.lower() and system.product_family in ["FILER", "E-SERIES", "V-SERIES"]:
                asup_off_count += 1
            if 'y' in system.pvr_flag.lower():
                eos_pvr_count += 1
            if system.response_profile not in response_profile_counts:
                response_profile_counts[system.response_profile] = 0
            response_profile_counts[system.response_profile] += 1

            # print_to_log("system.response_profile:" + system.response_profile + " response_profile_counts[system.response_profile]:" + str(response_profile_counts[system.response_profile]))

            # self.good_response_profiles = ["2HR PREMIUM ONSITE", "4HR PREMIUM ONSITE", "NBD PREMIUM ONSITE"]

            if self.good_response_profiles and system.response_profile not in self.good_response_profiles:
                bad_response_profile_list.append(system)

            if system.hw_service_level_status.lower() != "expired" and system.months_till_expire:
                if system.months_till_expire == 1 or system.months_till_expire == 2:
                    months_till_expire_count["Less than 3mos:"] += 1
                elif system.months_till_expire < 6 and system.months_till_expire >2:
                    months_till_expire_count["Less than 6mos:"] += 1
                elif system.months_till_expire < 12 and system.months_till_expire > 6:
                    months_till_expire_count["Less than 12mos:"] += 1

            if system.nrd_service_contract_id:
                nrd_contracts_count += 1
            if 'expired' in system.nrd_hw_service_level_status.lower():
                expired_nrd_count += 1
            if system.nrd_months_till_expire and system.nrd_months_till_expire < 3:
                nrd_months_till_expire_count["Less than 3mos:"] += 1
            elif system.nrd_months_till_expire and system.nrd_months_till_expire < 6:
                nrd_months_till_expire_count["Less than 6mos:"] += 1
            elif system.nrd_months_till_expire and system.nrd_months_till_expire < 12:
                nrd_months_till_expire_count["Less than 12mos:"] += 1

            for column_heading, attr_name, format_string in column_list:
                if column_heading not in self.column_width:
                    continue
                if column_heading in columns_to_exclude:
                    continue
                cell = ws.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                cell.number_format = format_string
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
                if attr_name == "hostname" and system.notes:
                    cell.comment = Comment(system.notes[-1], "NetApp")

                if attr_name == "controller_eos_date":
                    my_column += 1
                    try:
                        months_until_controller_eos = int((system.controller_eos_date - datetime.datetime.today()).days/30)
                        if months_until_controller_eos < 0:
                            months_until_eos_count["Controllers:"] += 1
                        elif months_until_controller_eos < 13:
                            months_until_eos_count["Controllers EOS next 12mos:"] += 1
                        elif months_until_controller_eos < 25:
                            months_until_eos_count["Controllers EOS 13-24mos:"] += 1
                        # print_to_log("Days till controller EOS: " + str(months_until_controller_eos))
                    except TypeError:
                        months_until_controller_eos = ''
                        # print_to_log("No controller EOS date")
                    cell = ws.cell(row=my_row, column=my_column, value=months_until_controller_eos)
                    cell.number_format = '0'
                    cell.font = self.data_cell_font
                    cell.alignment = self.data_cell_alignment
                    cell.border = self.data_cell_border
                    cell.fill = self.data_cell_fill

                if attr_name == "pvr_date":
                    my_column += 1
                    try:
                        months_until_pvr_end = int((system.pvr_date - datetime.datetime.today()).days/30)
                        # print_to_log("Days till controller EOS: " + str(months_until_pvr_end))
                    except TypeError:
                        months_until_pvr_end = ''
                        # print_to_log("No controller EOS date")
                    cell = ws.cell(row=my_row, column=my_column, value=months_until_pvr_end)
                    cell.number_format = '0'
                    cell.font = self.data_cell_font
                    cell.alignment = self.data_cell_alignment
                    cell.border = self.data_cell_border
                    cell.fill = self.data_cell_fill

                if attr_name == "first_eos_part":
                    my_column += 1
                    try:
                        months_until_first_eos = int((system.first_eos_date - datetime.datetime.today()).days/30)
                        # print_to_log("Days till controller EOS: " + str(months_until_first_eos))
                        if system.__getattribute__(attr_name) != "Controller":
                            if months_until_first_eos < 0:
                                months_until_eos_count["Other:"] += 1
                            elif months_until_first_eos < 13:
                                months_until_eos_count["Other EOS next 12mos:"] += 1
                            elif months_until_first_eos < 25:
                                months_until_eos_count["Other EOS 13-24mos:"] += 1

                    except TypeError:
                        months_until_first_eos = ''
                        # print_to_log("No controller EOS date")
                    cell = ws.cell(row=my_row, column=my_column, value=months_until_first_eos)
                    cell.number_format = '0'
                    cell.font = self.data_cell_font
                    cell.alignment = self.data_cell_alignment
                    cell.border = self.data_cell_border
                    cell.fill = self.data_cell_fill

                if attr_name == "partner_serial":
                    my_column += 1
                    if system.partner_serial and system.partner_serial in self.system_by_serial:
                        cell = ws.cell(row=my_row, column=my_column, value=self.system_by_serial[system.partner_serial].hostname)
                        # todo should we compare contracts and note differences here???
                    else:
                        cell = ws.cell(row=my_row, column=my_column, value='')
                    cell.number_format = '0'
                    cell.font = self.data_cell_font
                    cell.alignment = self.data_cell_alignment
                    cell.border = self.data_cell_border
                    cell.fill = self.data_cell_fill

                my_column += 1
            if system.notes:
                cell = ws.cell(row=my_row, column=my_column, value=system.notes[-1])
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
            else:
                cell = ws.cell(row=my_row, column=my_column, value='')
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
            my_row += 1

        # wb.save(filename=self.output_filename)

        month_ranges = [
            "Less than 3mos:",
            "Less than 6mos:",
            "Less than 12mos:"
        ]

        year_and_week = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W")

        with open("summary-" + year_and_week + ".txt", 'w', encoding='utf-8') as f:
            print("Hostnames/Serials in IB list:" + ',' + str(len(self.ib_hostnames)), file=f)
            print("Rows in IB report:" + ',' + str(ib_report_row_count), file=f)
            print("Serials processed in IB report:" + ',' + str(len(self.system_by_serial)), file=f)
            print("Serials skipped in IB report:" + ',' + str(self.serials_skipped), file=f)
            print("Serials with missing hostnames:" + ',' + str(len(self.unknown_hostname_systems)), file=f)
            print("Hostnames with missing groups:" + ',' + str(len(systems_missing_group)), file=f)
            print("Serials with expired HW Service:" + ',' + str(expired_count), file=f)
            print("Owner Companies:" + ',' + str(len(owner_companies)), file=f)
            print("Sites:" + ',' + str(len(sites)), file=f)
            print("ASUP Status=OFF:" + ',' + str(asup_off_count), file=f)
            print("EOS PVR Flag=Y:" + ',' + str(eos_pvr_count), file=f)
            print("NRD Contracts:" + ',' + str(nrd_contracts_count), file=f)
            print("Expired NRD Contracts:" + ',' + str(expired_nrd_count), file=f)

            for response_profile in HardwareContract.response_profiles_list:
                if response_profile in response_profile_counts:
                    if response_profile:
                        print(response_profile + ',' + str(response_profile_counts[response_profile]), file=f)
                    else:
                        print("NONE," + str(response_profile_counts[response_profile]), file=f)
                else:
                    print(response_profile + ',0', file=f)

            for months_till_expire in month_ranges:
                print(months_till_expire + ',' + str(months_till_expire_count[months_till_expire]), file=f)

            for months_till_expire in month_ranges:
                print("NRD " + months_till_expire + ',' + str(nrd_months_till_expire_count[months_till_expire]), file=f)

            for eos_key in eos_keys:
                print(eos_key + "," + str(months_until_eos_count[eos_key]), file=f)

        print_to_log("Generating summary...")

        summary = wb.create_sheet(title="Summary", index=0)
        summary['A1'] = "IB Team Installed Base Scrub v1.8 Summary"
        summary['A1'].font = self.title_font

        my_row = 3

        summary.cell(row=my_row-1, column=2, value="Week " + (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W")).font = Font(bold=True)

        summary.cell(row=my_row, column=1, value="Hostnames/Serials in IB list:")
        summary.cell(row=my_row, column=2, value=len(self.ib_hostnames)).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Rows in IB report:")
        summary.cell(row=my_row, column=2, value=ib_report_row_count).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Serials processed in IB report:")
        summary.cell(row=my_row, column=2, value=len(self.system_by_serial)).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Serials skipped in IB report:")
        summary.cell(row=my_row, column=2, value=self.serials_skipped).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Serials with missing hostnames:")
        summary.cell(row=my_row, column=2, value=len(self.unknown_hostname_systems)).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Hostnames with missing groups:")
        summary.cell(row=my_row, column=2, value=len(systems_missing_group)).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Serials with expired HW Service:")
        summary.cell(row=my_row, column=2, value=expired_count).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Owner Companies:")
        summary.cell(row=my_row, column=2, value=len(owner_companies)).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Sites:")
        summary.cell(row=my_row, column=2, value=len(sites)).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="ASUP Status=OFF:")
        summary.cell(row=my_row, column=2, value=asup_off_count).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="EOS PVR Flag=Y:")
        summary.cell(row=my_row, column=2, value=eos_pvr_count).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="NRD Contracts:")
        summary.cell(row=my_row, column=2, value=nrd_contracts_count).font = Font(bold=True)
        my_row += 1

        summary.cell(row=my_row, column=1, value="Expired NRD Contracts:")
        summary.cell(row=my_row, column=2, value=expired_nrd_count).font = Font(bold=True)
        my_row += 2

        summary.cell(row=my_row, column=1, value="Response Profiles:").font = Font(underline='single')
        my_row += 1

        for response_profile in HardwareContract.response_profiles_list:
            if response_profile in response_profile_counts:
                if response_profile:
                    summary.cell(row=my_row, column=1, value="   " + response_profile + ':')
                else:
                    summary.cell(row=my_row, column=1, value='   NONE:')
                summary.cell(row=my_row, column=2, value=response_profile_counts[response_profile]).font = Font(bold=True)
                my_row += 1
            else:
                if response_profile:
                    summary.cell(row=my_row, column=1, value="   " + response_profile + ':')
                else:
                    summary.cell(row=my_row, column=1, value='   NONE:')
                summary.cell(row=my_row, column=2, value=0).font = Font(bold=True)
                my_row += 1

        my_row += 1

        month_ranges = [
            "Less than 3mos:",
            "Less than 6mos:",
            "Less than 12mos:"
        ]

        summary.cell(row=my_row, column=1, value="Entitlements expiring in:").font = Font(underline='single')
        my_row += 1
        for months_till_expire in month_ranges:
            summary.cell(row=my_row, column=1, value="   " + months_till_expire)
            summary.cell(row=my_row, column=2, value=months_till_expire_count[months_till_expire]).font = Font(bold=True)
            my_row += 1
        for months_till_expire in month_ranges:
            summary.cell(row=my_row, column=1, value="   " + "NRD " + months_till_expire)
            summary.cell(row=my_row, column=2, value=nrd_months_till_expire_count[months_till_expire]).font = Font(bold=True)
            my_row += 1

        my_row += 1
        summary.cell(row=my_row, column=1, value="EOS systems:").font = Font(underline='single')
        my_row += 1
        for eos_key in eos_keys:
            summary.cell(row=my_row, column=1, value="   " + eos_key)
            summary.cell(row=my_row, column=2, value=months_until_eos_count[eos_key]).font = Font(bold=True)
            my_row += 1

        # todo add summary section for EOS details
        last_summary_row = my_row

        exception_column_list = [
            (serial_text, "serial", "0"),
            (hostname_text, "hostname", "General"),
            (hw_service_level_status_text, "hw_service_level_status", "General"),
            (response_profile_text, "response_profile", "General"),
            (product_family_text, "product_family", "General"),
            (platform_text, "platform", "General"),
            (os_version_text, "os_version", "General"),
            (account_name_text, "account_name", "General"),
            (site_text, "site", "General"),
            (group_text, "group", "General"),
            (entitlement_status_text, "entitlement_status", "General"),
            (shipped_date_text, "shipped_date", "mmm dd, yyyy"),
            (warranty_end_date_text, "warranty_end_date", "mmm dd, yyyy"),
            (contract_end_date_text, "contract_end_date", "mmm dd, yyyy"),
        ]

        my_row += 1

        with open("systems-" + (datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y%W"), 'wb') as f:
            pickle.dump(self.system_by_serial, f)

        # last_week = (datetime.date.today() + datetime.timedelta(days=1) - datetime.timedelta(days=7)).strftime("%Y%W")
        # check for history
        weeks = 1
        while weeks < 7:
            last_week = (datetime.date.today() + datetime.timedelta(days=1) - datetime.timedelta(days=7*weeks)).strftime("%Y%W")
            if not os.path.isfile("serials-" + last_week + ".txt"):
                weeks += 1
                continue
            else:
                print_to_log("Comparing to history from " + last_week)
                break

        previous_systems = {}
        if os.path.isfile("systems-" + last_week):
            print_to_log("Comparing to history from systems-" + last_week)
            with open("systems-" + last_week, 'rb') as f:
                previous_systems = pickle.load(f)

        # check for new and dropped systems
        # Betsy Added a row that tells you how many serials you have added
        if os.path.isfile("serials-" + last_week + ".txt"):
            print_to_log("Comparing to history from serials-" + last_week + ".txt")
            new_serials = set()
            dropped_serials = set()
            last_week_serials = set()
            with open("serials-" + last_week + ".txt", encoding='utf-8') as f:
                lines = f.readlines()
                for serial in lines:
                    last_week_serials.add(serial.strip())
            for serial in self.system_by_serial:
                if serial not in last_week_serials:
                    new_serials.add(serial)
            for serial in last_week_serials:
                if serial not in self.system_by_serial:
                    dropped_serials.add(serial)
            if new_serials:
                my_row += 1
                summary.cell(row=my_row, column=1, value="New systems:").font = Font(underline='single')
                summary.cell(row=my_row, column=2, value=len(new_serials)).font = Font(underline='single')
                my_row += 1
                my_column = 1
                for column_heading, attr_name, format_string in exception_column_list:
                    summary.cell(row=my_row, column=my_column, value=column_heading).font = Font(underline='single')
                    my_column += 1
                summary.cell(row=my_row, column=my_column, value="Notes").font = Font(underline='single')
                my_row += 1
                for serial in new_serials:
                    system = self.system_by_serial[serial]
                    my_column = 1
                    for column_heading, attr_name, format_string in exception_column_list:
                        cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                        cell.number_format = format_string
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                    if system.notes:
                        cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    else:
                        cell = summary.cell(row=my_row, column=my_column, value='')
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    my_row += 1
            if dropped_serials:
                my_row += 1
                summary.cell(row=my_row, column=1, value="Dropped systems:").font = Font(underline='single')
                summary.cell(row=my_row, column=2, value=len(dropped_serials)).font = Font(underline='single')
                my_row += 1
                my_column = 1
                for column_heading, attr_name, format_string in exception_column_list:
                    summary.cell(row=my_row, column=my_column, value=column_heading).font = Font(underline='single')
                    my_column += 1
                summary.cell(row=my_row, column=my_column, value="Notes").font = Font(underline='single')
                my_row += 1
                for serial in dropped_serials:
                    my_column = 1
                    if serial in previous_systems:
                        system = previous_systems[serial]
                        for column_heading, attr_name, format_string in exception_column_list:
                            cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                            cell.number_format = format_string
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                            my_column += 1
                        if system.notes:
                            cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                        else:
                            cell = summary.cell(row=my_row, column=my_column, value='')
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                    else:
                        cell = summary.cell(row=my_row, column=my_column, value=serial)
                        for column_heading, attr_name, format_string in exception_column_list:
                            cell = summary.cell(row=my_row, column=my_column)
                            cell.number_format = format_string
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                            my_column += 1
                    my_row += 1

        # check for owner company changes
        if os.path.isfile("owners-" + last_week + ".txt"):
            print_to_log("Comparing to history from owners-" + last_week + ".txt")
            new_owners = set()
            dropped_owners = set()
            last_week_owners = set()
            with open("owners-" + last_week + ".txt", encoding='utf-8') as f:
                lines = f.readlines()
                for owner in lines:
                    last_week_owners.add(owner.strip())
            for owner in self.set_of_owners:
                if owner not in last_week_owners:
                    new_owners.add(owner)
            for owner in last_week_owners:
                if owner not in self.set_of_owners:
                    dropped_owners.add(owner)
            if new_owners:
                my_row += 1
                summary.cell(row=my_row, column=1, value="New Owner Companies:").font = Font(underline='single')
                summary.cell(row=my_row, column=2, value="Number of serials:").font = Font(underline='single')
                my_row += 1
                for owner in new_owners:
                    summary.cell(row=my_row, column=1, value=owner)
                    summary.cell(row=my_row, column=2, value=count_by_owner[owner])
                    my_row += 1
            if dropped_owners:
                my_row += 1
                summary.cell(row=my_row, column=1, value="Dropped Owner Companies:").font = Font(underline='single')
                my_row += 1
                for owner in dropped_owners:
                    summary.cell(row=my_row, column=1, value=owner)
                    my_row += 1

        # check for site changes
        if os.path.isfile("sites-" + last_week + ".txt"):
            print_to_log("Comparing to history from sites-" + last_week + ".txt")
            new_sites = set()
            dropped_sites = set()
            last_week_sites = set()
            with open("sites-" + last_week + ".txt", encoding='utf-8') as f:
                lines = f.readlines()
                for site in lines:
                    last_week_sites.add(site.strip())
            for site in self.set_of_sites:
                if site not in last_week_sites:
                    new_sites.add(site)
            for site in last_week_sites:
                if site not in self.set_of_sites:
                    dropped_sites.add(site)
            if new_sites:
                my_row += 1
                summary.cell(row=my_row, column=1, value="New Sites:").font = Font(underline='single')
                summary.cell(row=my_row, column=2, value="Number of Serials:").font = Font(underline='single')
                my_row += 1
                for site in new_sites:
                    summary.cell(row=my_row, column=1, value=site)
                    summary.cell(row=my_row, column=2, value=count_by_site[site])
                    my_row += 1
            if dropped_sites:
                my_row += 1
                summary.cell(row=my_row, column=1, value="Dropped Sites:").font = Font(underline='single')
                my_row += 1
                for site in dropped_sites:
                    summary.cell(row=my_row, column=1, value=site)
                    my_row += 1

        # check for group changes
        if os.path.isfile("groups-" + last_week + ".txt"):
            print_to_log("Comparing to history from groups-" + last_week + ".txt")
            new_groups = set()
            dropped_groups = set()
            last_week_groups = set()
            with open("groups-" + last_week + ".txt", encoding='utf-8') as f:
                lines = f.readlines()
                for group in lines:
                    last_week_groups.add(group.strip())
            for group in self.set_of_groups:
                if group and group not in last_week_groups:
                    new_groups.add(group)
            for group in last_week_groups:
                if group not in self.set_of_groups:
                    dropped_groups.add(group)
            if new_groups:
                my_row += 1
                summary.cell(row=my_row, column=1, value="New Groups:").font = Font(underline='single')
                summary.cell(row=my_row, column=2, value="Number of Serials:").font = Font(underline='single')
                my_row += 1
                for group in new_groups:
                    summary.cell(row=my_row, column=1, value=group)
                    summary.cell(row=my_row, column=2, value=count_by_group[group])
                    my_row += 1
            if dropped_groups:
                my_row += 1
                summary.cell(row=my_row, column=1, value="Dropped Groups:").font = Font(underline='single')
                my_row += 1
                for group in dropped_groups:
                    summary.cell(row=my_row, column=1, value=group)
                    my_row += 1

        my_row += 1

        summary.cell(row=my_row, column=1, value="Service Exceptions:").font = Font(underline='single')
        my_row += 1
        my_column = 1
        for column_heading, attr_name, format_string in exception_column_list:
            summary.cell(row=my_row, column=my_column, value=column_heading).font = Font(underline='single')
            my_column += 1
        summary.cell(row=my_row, column=my_column, value="Notes").font = Font(underline='single')
        my_row += 1

        exception_list = []
        for system in expired_list:
            if system.notes:
                if len(system.notes[-1]) > 5 and "GREEN:" in system.notes[-1]:
                    # don't display GREEN exceptions
                    continue

            my_column = 1
            for column_heading, attr_name, format_string in exception_column_list:
                cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                cell.number_format = format_string
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
                my_column += 1
            if system.notes:
                cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
            else:
                cell = summary.cell(row=my_row, column=my_column, value='')
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill

            exception_list.append(system)
            my_row += 1
        for system in bad_response_profile_list:
            if system not in expired_list:
                if system.notes:
                    if len(system.notes[-1]) > 5 and "GREEN:" in system.notes[-1]:
                        # don't display GREEN exceptions
                        continue
            my_column = 1
            for column_heading, attr_name, format_string in exception_column_list:
                cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                cell.number_format = format_string
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
                my_column += 1
            if system.notes:
                cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
            else:
                cell = summary.cell(row=my_row, column=my_column, value='')
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
                my_row += 1
            exception_list.append(system)

        for system in self.list_of_systems:
            if system.notes:
                if ("RED:" in system.notes[-1] or "YELLOW:" in system.notes[-1]) and system not in exception_list:
                    my_column = 1
                    for column_heading, attr_name, format_string in exception_column_list:
                        cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                        cell.number_format = format_string
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                    if system.notes:
                        cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    else:
                        cell = summary.cell(row=my_row, column=my_column, value='')
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    my_row += 1
        my_row += 1

        summary.cell(row=my_row, column=1, value="Missing hostnames:").font = Font(underline='single')
        my_row += 1
        my_column = 1
        for column_heading, attr_name, format_string in exception_column_list:
            summary.cell(row=my_row, column=my_column, value=column_heading).font = Font(underline='single')
            my_column += 1
        summary.cell(row=my_row, column=my_column, value="Notes").font = Font(underline='single')
        my_row += 1
        for serial in self.unknown_hostname_systems:
            # should we display switches here?
            if debug_it:
                print_to_log("Missing hostname for serial: " + serial)
            system = self.unknown_hostname_systems[serial]
            my_column = 1
            for column_heading, attr_name, format_string in exception_column_list:
                cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                cell.number_format = format_string
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
                my_column += 1
            if system.notes:
                cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
            else:
                cell = summary.cell(row=my_row, column=my_column, value='')
                cell.font = self.data_cell_font
                cell.alignment = self.data_cell_alignment
                cell.border = self.data_cell_border
                cell.fill = self.data_cell_fill
            my_row += 1

        my_row += 1
        if self.ib_hostnames:
            summary.cell(row=my_row, column=1, value="Hostnames/Serials missing from IB Details report:").font = Font(underline='single')
            my_row += 1
            no_missing_systems = True
            # todo use serials instead
            for hostname in self.ib_hostnames:
                if hostname not in self.system_by_hostname and hostname not in self.system_by_serial:
                    if debug_it:
                        print_to_log("Hostname/Serial missing from IB Details report: " + hostname)
                    summary.cell(row=my_row, column=1, value=hostname)
                    no_missing_systems = False
                    my_row += 1
            if no_missing_systems:
                summary.cell(row=my_row, column=1, value="None")
                my_row += 1
            my_row += 1

        if self.check_groups:
            summary.cell(row=my_row, column=1, value="Hostnames with missing groups:").font = Font(underline='single')
            my_row += 1
            my_column = 1
            for column_heading, attr_name, format_string in exception_column_list:
                summary.cell(row=my_row, column=my_column, value=column_heading).font = Font(underline='single')
                my_column += 1
            summary.cell(row=my_row, column=my_column, value="Notes").font = Font(underline='single')
            my_row += 1
            if len(systems_missing_group) > 0:
                # todo use serials instead
                for system in systems_missing_group:
                    my_column = 1
                    for column_heading, attr_name, format_string in exception_column_list:
                        cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                        cell.number_format = format_string
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                    if system.notes:
                        cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    else:
                        cell = summary.cell(row=my_row, column=my_column, value='')
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    my_row += 1
            else:
                summary.cell(row=my_row, column=1, value="None")
                my_row += 1

        if self.ib_hostnames:
            my_row += 1
            summary.cell(row=my_row, column=1, value="Systems not in IB List:").font = Font(underline='single')
            my_row += 1
            my_column = 1
            for column_heading, attr_name, format_string in exception_column_list:
                summary.cell(row=my_row, column=my_column, value=column_heading).font = Font(underline='single')
                my_column += 1
            summary.cell(row=my_row, column=my_column, value="Notes").font = Font(underline='single')
            my_row += 1
            # todo use system by serial instead
            for hostname, system in self.system_by_hostname.items():
                if hostname not in self.ib_hostnames and system.serial not in self.ib_hostnames:
                    if debug_it:
                        print_to_log("Hostname not in ib list: " + hostname)
                    my_column = 1
                    for column_heading, attr_name, format_string in exception_column_list:
                        cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                        cell.number_format = format_string
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                    if system.notes:
                        cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    else:
                        cell = summary.cell(row=my_row, column=my_column, value='')
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    my_row += 1
                else:
                    if debug_it:
                        print_to_log(','.join([hostname, system.serial, system.entitlement_status, system.sc_entitlement_status]))
                    if system.entitlement_status != system.sc_entitlement_status:
                        print_to_log("Entitlement status mismatch " + system.serial)
            for serial, system in self.unknown_hostname_systems.items():
                if serial not in self.ib_hostnames:
                    if debug_it:
                        print_to_log("Serial not in ib list: " + hostname)
                    my_column = 1
                    for column_heading, attr_name, format_string in exception_column_list:
                        cell = summary.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                        cell.number_format = format_string
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                    if system.notes:
                        cell = summary.cell(row=my_row, column=my_column, value=system.notes[-1])
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    else:
                        cell = summary.cell(row=my_row, column=my_column, value='')
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                    my_row += 1
                else:
                    if debug_it:
                        print_to_log(','.join([hostname, system.serial, system.entitlement_status, system.sc_entitlement_status]))
                    if system.entitlement_status != system.sc_entitlement_status:
                        print_to_log("Entitlement status mismatch " + system.serial)

        summary.column_dimensions['A'].width = 36
        summary.column_dimensions['B'].width = 24
        summary.column_dimensions['C'].width = 24
        summary.column_dimensions['D'].width = 24
        summary.column_dimensions['E'].width = 24
        summary.column_dimensions['F'].width = 24
        summary.column_dimensions['G'].width = 24
        summary.column_dimensions['H'].width = 24
        summary.column_dimensions['I'].width = 24
        summary.column_dimensions['J'].width = 24
        summary.column_dimensions['K'].width = 24
        summary.column_dimensions['L'].width = 24
        summary.column_dimensions['M'].width = 24
        summary.column_dimensions['N'].width = 24
        summary.column_dimensions['O'].width = 24
        # wb.save(filename=self.output_filename)

        # Betsy add - Group Name/Product Family
        print_to_log("Checking for changes...")
        my_row += 1
        summary.cell(row=my_row, column=1, value="Changes from previous report:").font = Font(underline='single')
        my_row += 1
        my_column = 1
        summary.cell(row=my_row, column=my_column, value="Serial Number").font = Font(underline='single')
        my_column += 1
        summary.cell(row=my_row, column=my_column, value="Hostname").font = Font(underline='single')
        my_column += 1
        summary.cell(row=my_row, column=my_column, value="Product Family").font = Font(underline='single')
        my_column += 1
        summary.cell(row=my_row, column=my_column, value="Group Name").font = Font(underline='single')
        my_column += 1
        summary.cell(row=my_row, column=my_column, value="Field").font = Font(underline='single')
        my_column += 1
        summary.cell(row=my_row, column=my_column, value="Old Value").font = Font(underline='single')
        my_column += 1
        summary.cell(row=my_row, column=my_column, value="New Value").font = Font(underline='single')
        my_row += 1
        count_of_systems = len(self.list_of_systems)
        count = 0
        for system in self.list_of_systems:
            count += 1
            if count % 1000 == 0:
                print_to_log("Checked " + str(count) + " systems")
            if system.serial in previous_systems:
                changes = system.list_changes(previous_systems[system.serial])
                for k, new_value, old_value in changes:
                    if k in attribute_text:
                        # changes_text = ' '.join([system.serial, system.hostname, attribute_text[k], "changed from", str(old_value), "to",str(new_value)])
                        # print_to_log(changes_text)
                        # summary.cell(row=my_row, column=1, value=changes_text)
                        my_column = 1
                        cell = summary.cell(row=my_row, column=my_column, value=system.serial)
                        cell.number_format = "0"
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                        cell = summary.cell(row=my_row, column=my_column, value=system.hostname)
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                        cell = summary.cell(row=my_row, column=my_column, value=system.product_family)
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                        cell = summary.cell(row=my_row, column=my_column, value=system.group)
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                        cell = summary.cell(row=my_row, column=my_column, value=attribute_text[k])
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                        cell = summary.cell(row=my_row, column=my_column, value=str(old_value))
                        cell.number_format = attribute_format_string[k]
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill
                        my_column += 1
                        cell = summary.cell(row=my_row, column=my_column, value=str(new_value))
                        cell.number_format = attribute_format_string[k]
                        cell.font = self.data_cell_font
                        cell.alignment = self.data_cell_alignment
                        cell.border = self.data_cell_border
                        cell.fill = self.data_cell_fill

                        my_row += 1

        # check for history
        weeks = 1
        week_count = 0
        missing_weeks = 0
        my_column = 3
        while week_count < 6 and missing_weeks < 52:

            year_and_week = (datetime.date.today() + datetime.timedelta(days=1) - datetime.timedelta(days=7*weeks)).strftime("%Y%W")
            if not os.path.isfile("summary-" + year_and_week + ".txt"):
                weeks += 1
                missing_weeks += 1
                continue

            with open("summary-" + year_and_week + ".txt", encoding='utf-8') as f:
                week_count += 1
                lines = f.readlines()
                history = {}
                history["Hostnames/Serials in IB list:"] = 0
                history["Rows in IB report:"] = 0
                history["Serials processed in IB report:"] = 0
                history["Serials skipped in IB report:"] = 0
                history["Serials with missing hostnames:"] = 0
                history["Serials with expired HW Service:"] = 0
                history["Owner Companies:"] = 0
                history["Sites:"] = 0
                history["ASUP Status=OFF:"] = 0
                history["EOS PVR Flag=Y:"] = 0
                history["2HR PREMIUM ONSITE:"] = 0
                history["4HR PREMIUM ONSITE:"] = 0
                history["4HR PARTS REPLACE:"] = 0
                history["NBD PREMIUM ONSITE:"] = 0
                history["NBD PARTS REPLACE:"] = 0
                history["NBD PARTS DELIVERY:"] = 0
                history["NONE:"] = 0
                history["Less than 3mos:"] = 0
                history["Less than 6mos:"] = 0
                history["Less than 12mos:"] = 0
                history["NRD Contracts:"] = 0
                history["Expired NRD Contracts:"] = 0
                history["NRD Less than 3mos:"] = 0
                history["NRD Less than 6mos:"] = 0
                history["NRD Less than 12mos:"] = 0
                history["Hostnames with missing groups:"] = 0
                history["Controllers:"] = 0
                history["Controllers EOS next 12mos:"] = 0
                history["Controllers EOS 13-24mos:"] = 0
                history["Other:"] = 0
                history["Other EOS next 12mos:"] = 0
                history["Other EOS 13-24mos:"] = 0
                history[""] = 0
                for line in lines:
                    temp_words = line.split(",", 1)
                    if len(temp_words) == 2:
                        history[temp_words[0]] = temp_words[1]
                my_row = 3
                summary.cell(row=my_row - 1, column=my_column, value="Week " + year_and_week)

                summary.cell(row=my_row, column=my_column, value=int(history.get("Hostnames/Serials in IB list:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Rows in IB report:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Rows in IB report:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Serials processed in IB report:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Serials processed in IB report:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Serials skipped in IB report:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Serials skipped in IB report:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Serials with missing hostnames:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Serials with missing hostnames:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Hostnames with missing groups:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Hostnames with missing groups:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Serials with expired HW Service:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Serials with expired HW Service:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Owner Companies:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Owner Companies:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Sites:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Sites:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="ASUP Status=OFF:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("ASUP Status=OFF:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="EOS PVR Flag=Y:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("EOS PVR Flag=Y:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="NRD Contracts:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("NRD Contracts:", "0")))
                my_row += 1

                summary.cell(row=my_row, column=1, value="Expired NRD Contracts:")
                summary.cell(row=my_row, column=my_column, value=int(history.get("Expired NRD Contracts:", "0")))
                my_row += 3

                for response_profile in HardwareContract.response_profiles_list:
                    if response_profile:
                        summary.cell(row=my_row, column=my_column, value=int(history.get(response_profile, "0")))
                    else:
                        summary.cell(row=my_row, column=my_column, value=int(history.get("NONE", "0")))
                    my_row += 1
                my_row += 2

                for months_till_expire in month_ranges:
                    summary.cell(row=my_row, column=my_column, value=int(history.get(months_till_expire, "0")))
                    my_row += 1
                for months_till_expire in month_ranges:
                    summary.cell(row=my_row, column=my_column, value=int(history.get("NRD " + months_till_expire, "0")))
                    my_row += 1
                my_row += 2
                for eos_key in eos_keys:
                    summary.cell(row=my_row, column=my_column, value=int(history.get(eos_key, "0")))
                    my_row += 1

                # summary.column_dimensions[get_column_letter(my_column)].width = 15
                my_column += 1
                weeks += 1

        for row_iter in range(3, last_summary_row):
            for col_iter in range(2, my_column):
                summary.cell(row=row_iter, column=col_iter).alignment = self.data_cell_alignment
                summary.cell(row=row_iter, column=col_iter).border = self.data_cell_border
                summary.cell(row=row_iter, column=col_iter).fill = self.data_cell_fill

        # wb.save(filename=self.output_filename)
        flash_file = "flash.xlsx"
        if os.path.isfile(flash_file):
            print_to_log("Reading flash file")
            try:
                flash_wb= load_workbook(flash_file, read_only=True)
                need_flash_asups = False
                sheet = flash_wb.active
                rows = sheet.rows
                header_row = [cell.value for cell in next(rows)]
                while header_row[0] != "Serial Number":
                    header_row = [cell.value for cell in next(rows)]
                data = []
                count = 0
                for row in rows:
                    count += 1
                    if count % 1000 == 0:
                        print_to_log("Current row: " + str(count))
                    record = {}
                    for key, cell in zip(header_row, row):
                        record[key] = cell.value
                    data.append(record)
                flash_rows = len(data)
                for row in data:
                    if debug_it:
                        print_to_log("Parsing row " + str(my_row) + " from flash file")
                    if not row[serial_text]:
                        continue
                    try:
                        serial = str(int(row[serial_text]))
                    except ValueError:
                        serial = row[serial_text]
                    if not serial:
                        print_to_log("Blank serial on report row " + str(my_row))
                        my_row += 1
                        continue

                    if serial in self.serials_to_ignore:
                        if debug_it:
                            print_to_log("Ignore serial " + serial)
                        my_row += 1
                        continue

                    if serial in self.system_by_serial:
                        system = self.system_by_serial[serial]
                        try:
                            flash_serial = str(int(row["Flash Serial"]))
                        except ValueError:
                            flash_serial = row["Flash Serial"]
                        system.flash_list.append((flash_serial, row["Flash Type"], row["Flash Model"], row["Flash Info"]))

                    my_row += 1

            except:
                print_to_log("Problem reading flash file, please try downloading a new version before running IB Scrub again")
                return
            flash_tab = wb.create_sheet(title="Flash")
            flash_tab['A1'] = "FlashCache and PAM Details from AutoSupports"
            flash_tab['A1'].font = self.title_font

            column_list = [
                (serial_text, "serial", "0"),
                (company_text, "owner", "General"),
                (group_text, "group", "General"),
                (hostname_text, "hostname", "General"),
                (site_text, "site", "General"),
                (os_version_text, "os_version", "General"),
                (product_family_text, "product_family", "General"),
                (platform_text, "platform", "General"),
                (service_level_text, "service_level", "General"),
                (entitlement_status_text, "entitlement_status", "General"),
                (warranty_end_date_text, "warranty_end_date", "mmm dd, yyyy"),
                (service_type_text, "service_type", "General"),
                (response_profile_text, "response_profile", "General"),
                (service_contract_id_text, "service_contract_id", "General"),
                (contract_end_date_text, "contract_end_date", "mmm dd, yyyy"),
                (months_till_expire_text, "months_till_expire", "0"),
                (hw_service_level_status_text, "hw_service_level_status", "General")
            ]

            flash_columns = ["service_level", "entitlement_status", "warranty_end_date", "service_type", "response_profile", "service_contract_id", "contract_end_date", "months_till_expire", "hw_service_level_status", "site"]

            my_column = 1
            my_row = 2
            for column_heading, attr_name, format_string in column_list:
                cell = flash_tab.cell(row=my_row, column=my_column, value=column_heading)
                cell.font = self.heading_font
                cell.alignment = self.heading_alignment
                cell.border = self.heading_border
                cell.fill = self.heading_fill
                flash_tab.column_dimensions[get_column_letter(my_column)].width = self.column_width[column_heading]
                my_column += 1
                if attr_name in flash_columns:
                    cell = flash_tab.cell(row=my_row, column=my_column, value="Flash " + column_heading)
                    cell.font = self.heading_font
                    cell.alignment = self.heading_alignment
                    cell.border = self.heading_border
                    cell.fill = self.heading_fill
                    flash_tab.column_dimensions[get_column_letter(my_column)].width = self.column_width[column_heading]
                    my_column += 1

            cell = flash_tab.cell(row=my_row, column=my_column, value="Flash Serial")
            cell.font = self.heading_font
            cell.alignment = self.heading_alignment
            cell.border = self.heading_border
            cell.fill = self.heading_fill
            flash_tab.column_dimensions[get_column_letter(my_column)].width = 10
            my_column += 1
            cell = flash_tab.cell(row=my_row, column=my_column, value="Flash Type")
            cell.font = self.heading_font
            cell.alignment = self.heading_alignment
            cell.border = self.heading_border
            cell.fill = self.heading_fill
            flash_tab.column_dimensions[get_column_letter(my_column)].width = 10
            my_column += 1
            cell = flash_tab.cell(row=my_row, column=my_column, value="Flash Model")
            cell.font = self.heading_font
            cell.alignment = self.heading_alignment
            cell.border = self.heading_border
            cell.fill = self.heading_fill
            flash_tab.column_dimensions[get_column_letter(my_column)].width = 10
            my_column += 1
            cell = flash_tab.cell(row=my_row, column=my_column, value="Flash Info")
            cell.font = self.heading_font
            cell.alignment = self.heading_alignment
            cell.border = self.heading_border
            cell.fill = self.heading_fill
            flash_tab.column_dimensions[get_column_letter(my_column)].width = 25
            my_column += 1

            my_row = 3

            sn_list = list(self.system_by_serial.keys())

            for system in self.list_of_systems:

                if my_row % 100 == 0:
                    print_to_log(str(my_row) + " serials of " + str(flash_rows) + " processed...")
                if system.product_family in ["FILER", "V-SERIES"]:
                    if system.flash_list:
                        for flash_serial, flash_type, flash_model, flash_info in system.flash_list:
                            # print_to_log(flash_serial + ',' + flash_type + ',' + flash_model + ',' + flash_info)
                            my_column = 1
                            for column_heading, attr_name, format_string in column_list:
                                if attr_name in flash_columns:
                                    cell = flash_tab.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                                    cell.number_format = format_string
                                    cell.font = self.data_cell_font
                                    cell.alignment = self.data_cell_alignment
                                    cell.border = self.data_cell_border
                                    cell.fill = self.data_cell_fill
                                    my_column += 1
                                    if flash_serial in self.flash_by_serial:
                                        cell = flash_tab.cell(row=my_row, column=my_column, value=self.flash_by_serial[flash_serial].__getattribute__(attr_name))
                                    else:
                                        cell = flash_tab.cell(row=my_row, column=my_column, value='')
                                else:
                                    cell = flash_tab.cell(row=my_row, column=my_column, value=system.__getattribute__(attr_name))
                                cell.number_format = format_string
                                cell.font = self.data_cell_font
                                cell.alignment = self.data_cell_alignment
                                cell.border = self.data_cell_border
                                cell.fill = self.data_cell_fill
                                my_column += 1
                            cell = flash_tab.cell(row=my_row, column=my_column, value=flash_serial)
                            cell.number_format = "0"
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                            my_column += 1
                            cell = flash_tab.cell(row=my_row, column=my_column, value=flash_type)
                            cell.number_format = "0"
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                            my_column += 1
                            cell = flash_tab.cell(row=my_row, column=my_column, value=flash_model)
                            cell.number_format = "0"
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                            my_column += 1
                            cell = flash_tab.cell(row=my_row, column=my_column, value=flash_info)
                            cell.number_format = "0"
                            cell.font = self.data_cell_font
                            cell.alignment = self.data_cell_alignment
                            cell.border = self.data_cell_border
                            cell.fill = self.data_cell_fill
                            my_row += 1
        else:
            print_to_log("Missing flash file please use naming convention flash.xlsx or download IB Scrub Flash Details report from http://sam-reporting.hq.netapp.com/reports/asupcheck/")
        wb.save(filename=self.output_filename)

        print_to_log("Finished generating " + self.output_filename)
        NetAppSystem.failed_attributes = []
        time.sleep(6)
        if platform.system() == "Darwin":
            # mac
            os.system("open " + self.output_filename)
        else:
            # pc
            os.system("start " + self.output_filename)

    def get_ib_serial_count(self):
        return len(self.system_by_serial)

    def get_ib_report_name(self):
        return self.output_filename

    def map_hostnames(self):
        try:
            with open(missing_hostname_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()
            for line in lines:
                if line.strip() and line[0] != '#':
                    temp_words = line.strip().split(',')
                    if len(temp_words) == 2:
                        self.hostname_from_serial[temp_words[0]] = temp_words[1].strip().lower()
                        continue
                    temp_words = line.strip().split()
                    if len(temp_words) == 2:
                        self.hostname_from_serial[temp_words[0]] = temp_words[1].strip().lower()

        except FileNotFoundError:
            with open(missing_hostname_filename, 'w', encoding='utf-8') as f:
                print("#use the following format (without the #)", file=f)
                print("#comments denoted by line starting with '#'", file=f)
                print("#serial,hostname or serial hostname", file=f)
                print_to_log("Created " + missing_hostname_filename)

    def process_companies_to_ignore(self):
        try:
            with open(companies_to_ignore_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()

            for line in lines:
                if line.strip() and line[0] != '#':
                    self.companies_to_ignore.add(line.strip())
                    print_to_log("Company to ignore: " + line.strip())
        except FileNotFoundError:
            with open(companies_to_ignore_filename, 'w', encoding='utf-8') as f:
                print("#list companies to skip, one per line, case sensitive", file=f)
                print_to_log("Created " + companies_to_ignore_filename)

    def process_serials_to_ignore(self):
        try:
            with open(serials_to_ignore_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()

            for line in lines:
                if line.strip() and line[0] != '#':
                    self.serials_to_ignore.add(line.strip())
                    print_to_log("Serial to ignore: " + line.strip())
        except FileNotFoundError:
            with open(serials_to_ignore_filename, 'w', encoding='utf-8') as f:
                print("#list serials to skip, one per line", file=f)
                print_to_log("Created " + serials_to_ignore_filename)

    def process_hostnames_to_ignore(self):
        try:
            with open(hostnames_to_ignore_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()

            for line in lines:
                if line.strip() and line[0] != '#':
                    self.hostnames_to_ignore.add(line.strip().lower())
                    print_to_log("Serial to ignore: " + line.strip().lower())
        except FileNotFoundError:
            with open(hostnames_to_ignore_filename, 'w', encoding='utf-8') as f:
                print("#list hostnames to skip, one per line", file=f)
                print_to_log("Created " + hostnames_to_ignore_filename)

    def process_groups_to_ignore(self):
        try:
            with open(groups_to_ignore_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()

            for line in lines:
                if line.strip() and line[0] != '#':
                    self.groups_to_ignore.add(line.strip())
                    print_to_log("Group to ignore: " + line.strip())
        except FileNotFoundError:
            with open(groups_to_ignore_filename, 'w', encoding='utf-8') as f:
                print("#list groups to skip, one per line, case sensitive (use uppercase)", file=f)
                print_to_log("Created " + groups_to_ignore_filename)

    def process_sites_to_ignore(self):
        try:
            with open(sites_to_ignore_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()

            for line in lines:
                if line.strip() and line[0] != '#':
                    self.sites_to_ignore.add(line.strip())
                    print_to_log("Site to ignore: " + line.strip())
        except FileNotFoundError:
            with open(sites_to_ignore_filename, 'w', encoding='utf-8') as f:
                print("#list sites to ignore, one per line, case sensitive (use uppercase)", file=f)
                print_to_log("Created " + sites_to_ignore_filename)

    def process_ib_list(self):
        try:
            with open(installed_base_hostname_list_filename, encoding='utf-8') as f:
                lines = f.read().splitlines()

            use_ib_hostnames = True
            self.check_groups = False
            for line in lines:
                if line.strip():
                    hostname = line.strip().lower()
                    if "#option:ignore_ib.txt" in hostname:
                        print_to_log("#option:ignore_ib.txt - skipping ib.txt checks")
                        self.ib_hostnames.clear()
                        use_ib_hostnames = False
                        # self.check_groups
                    elif "#option:check_groups:true" in line.lower():
                        self.check_groups = True
                        print_to_log(line.strip())
                    elif use_ib_hostnames and hostname[0] != '#':
                        if hostname in self.ib_hostnames:
                            print_to_log("Warning: skipping duplicate " + hostname)
                        else:
                            self.ib_hostnames.add(hostname)
                    elif "#option:customer_list:" in line:
                        temp_words = line.strip().split(':')
                        if len(temp_words) == 3:
                            self.customer_id_list = temp_words[2].split(',')
                            print_to_log(line.strip())
                    elif "#option:good_response_profiles:" in line:
                        temp_words = line.strip().split(':')
                        if len(temp_words) == 3:
                            if ',' in temp_words[2]:
                                self.good_response_profiles = temp_words[2].split(',')
                            elif temp_words[2]:
                                self.good_response_profiles = [temp_words[2]]
                            print_to_log(line.strip())

        except FileNotFoundError:
            with open(installed_base_hostname_list_filename, 'w', encoding='utf-8') as f:
                print("#option:ignore_ib.txt", file=f)
                print("#remove the line above or add an extra '#' to enable ib checks", file=f)
                print("#list hostnames in installed base, one per line", file=f)
                print("#option:customer_list:", file=f)
                print("#option:good_response_profiles:", file=f)
                print("#option:check_groups:False")
                # self.good_response_profiles = "2HR PREMIUM ONSITE,4HR PREMIUM ONSITE,NBD PREMIUM ONSITE".split(',')
                self.customer_id_list = []
                self.check_groups = False
                print_to_log("Created " + installed_base_hostname_list_filename)

        if self.good_response_profiles:
            print_to_log("Expected response profiles: " + ' '.join(self.good_response_profiles))
